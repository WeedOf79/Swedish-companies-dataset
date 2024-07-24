# Valerio Malerba, Uppsala, 2023-2024

# This script reads raw data in form of CSVs and clean them, creating a dataset that can be suited for machine learning and data analysis.

import pandas as pd
import os
from openpyxl import load_workbook
import datetime
from geopy.geocoders import Nominatim
import numpy as np

# CONSTANTS

DATA_DIRECTORY = 'D:\Documents\Python Scripts\Scrapers\Bolagsskrapare\Raw data' # Folder containing the CSV files
EARLIEST_YEAR = 2012
FOLDER = "..\\"
RAW_DATASET_FILENAME = "Swedish Ltd companies dataset 2016-2022.csv"
CLEAN_DATASET_FILENAME = "Clean_data.csv"
COORDINATES_FILENAME = "Swedish_location_coordinates.csv"
EXPORT_CATEGORY_MAP = 'Category_map.csv'
EXPORT_REGION_MAP = 'Region_map.csv'
ENCODING = 'utf-16'
MAX_ROWS = -1  # Use -1 for all rows
N_LAT_BINS = 14
N_LONG_BINS = 6
EXPORT = False




# INIT

# List of features that have too many NaNs
# The features that are included in the following list have been determined by means of this formula:
# (Clean_data.isna().sum() * 100 / 446186).astype(int) > 50
# The formula is applied to the full dataset (446186 rows). Those features which had more than half of years marked with True, have been made eligible for deletion.
Unrelevant_features = [
    "Net revenue per employee_2012",
    "Net revenue per employee_2013",
    "Net revenue per employee_2014",
    "Net revenue per employee_2015",
    "Net revenue per employee_2016",
    "Net revenue per employee_2017",
    "Net revenue per employee_2018",
    "Net revenue per employee_2019",
    "Net revenue per employee_2020",
    "Net revenue per employee_2021",
    "Net revenue per employee_2022",
    # "Personnel costs per employee_2012",
    # "Personnel costs per employee_2013",
    # "Personnel costs per employee_2014",
    # "Personnel costs per employee_2015",
    # "Personnel costs per employee_2016",
    # "Personnel costs per employee_2017",
    # "Personnel costs per employee_2018",
    # "Personnel costs per employee_2019",
    # "Personnel costs per employee_2020",
    # "Personnel costs per employee_2021",
    # "Personnel costs per employee_2022",
    # "Salaries for other employees_2012",
    # "Salaries for other employees_2013",
    # "Salaries for other employees_2014",
    # "Salaries for other employees_2015",
    # "Salaries for other employees_2016",
    # "Salaries for other employees_2017",
    # "Salaries for other employees_2018",
    # "Salaries for other employees_2019",
    # "Salaries for other employees_2020",
    # "Salaries for other employees_2021",
    # "Salaries for other employees_2022",
    "Of which bonus salary to other employees_2012",
    "Of which bonus salary to other employees_2013",
    "Of which bonus salary to other employees_2014",
    "Of which bonus salary to other employees_2015",
    "Of which bonus salary to other employees_2016",
    "Of which bonus salary to other employees_2017",
    "Of which bonus salary to other employees_2018",
    "Of which bonus salary to other employees_2019",
    "Of which bonus salary to other employees_2020",
    "Of which bonus salary to other employees_2021",
    "Of which bonus salary to other employees_2022",
    "Salaries for board & CEO_2012",
    "Salaries for board & CEO_2013",
    "Salaries for board & CEO_2014",
    "Salaries for board & CEO_2015",
    "Salaries for board & CEO_2016",
    "Salaries for board & CEO_2017",
    "Salaries for board & CEO_2018",
    "Salaries for board & CEO_2019",
    "Salaries for board & CEO_2020",
    "Salaries for board & CEO_2021",
    "Salaries for board & CEO_2022",
    "Of which bonus to board & CEO_2012",
    "Of which bonus to board & CEO_2013",
    "Of which bonus to board & CEO_2014",
    "Of which bonus to board & CEO_2015",
    "Of which bonus to board & CEO_2016",
    "Of which bonus to board & CEO_2017",
    "Of which bonus to board & CEO_2018",
    "Of which bonus to board & CEO_2019",
    "Of which bonus to board & CEO_2020",
    "Of which bonus to board & CEO_2021",
    "Of which bonus to board & CEO_2022",
    "Minority interests_2015",
    "Minority interests_2016",
    "Minority interests_2017",
    "Minority interests_2018",
    "Minority interests_2020",
    # "Social costs_2012",
    # "Social costs_2013",
    # "Social costs_2014",
    # "Social costs_2015",
    # "Social costs_2016",
    # "Social costs_2017",
    # "Social costs_2018",
    # "Social costs_2019",
    # "Social costs_2020",
    # "Social costs_2021",
    # "Social costs_2022"
    ]

# This table contains both sorted columns and corresponding translations in english
Sorted_columns_translation_table = {
    "jurnamn": "juridical name",
    "orgnr": "organization number",
    "abv_ugrupp": "category",
    "ba_postort": "location",
    "Status": "Status",
    "Bolaget registrerat": "registration year",
    "Ägandeförhållande": "ownership",
    "Kommunsäte": "municipal seat",
    "Antal_anställda_2012": "Number of employees_2012",
    "Antal_anställda_2013": "Number of employees_2013",
    "Antal_anställda_2014": "Number of employees_2014",
    "Antal_anställda_2015": "Number of employees_2015",
    "Antal_anställda_2016": "Number of employees_2016",
    "Antal_anställda_2017": "Number of employees_2017",
    "Antal_anställda_2018": "Number of employees_2018",
    "Antal_anställda_2019": "Number of employees_2019",
    "Antal_anställda_2020": "Number of employees_2020",
    "Antal_anställda_2021": "Number of employees_2021",
    "Antal_anställda_2022": "Number of employees_2022",
    "Nettoomsättning per anställd (tkr)_2012": "Net revenue per employee_2012",
    "Nettoomsättning per anställd (tkr)_2013": "Net revenue per employee_2013",
    "Nettoomsättning per anställd (tkr)_2014": "Net revenue per employee_2014",
    "Nettoomsättning per anställd (tkr)_2015": "Net revenue per employee_2015",
    "Nettoomsättning per anställd (tkr)_2016": "Net revenue per employee_2016",
    "Nettoomsättning per anställd (tkr)_2017": "Net revenue per employee_2017",
    "Nettoomsättning per anställd (tkr)_2018": "Net revenue per employee_2018",
    "Nettoomsättning per anställd (tkr)_2019": "Net revenue per employee_2019",
    "Nettoomsättning per anställd (tkr)_2020": "Net revenue per employee_2020",
    "Nettoomsättning per anställd (tkr)_2021": "Net revenue per employee_2021",
    "Nettoomsättning per anställd (tkr)_2022": "Net revenue per employee_2022",
    "Personalkostnader per anställd (tkr)_2012": "Personnel costs per employee_2012",
    "Personalkostnader per anställd (tkr)_2013": "Personnel costs per employee_2013",
    "Personalkostnader per anställd (tkr)_2014": "Personnel costs per employee_2014",
    "Personalkostnader per anställd (tkr)_2015": "Personnel costs per employee_2015",
    "Personalkostnader per anställd (tkr)_2016": "Personnel costs per employee_2016",
    "Personalkostnader per anställd (tkr)_2017": "Personnel costs per employee_2017",
    "Personalkostnader per anställd (tkr)_2018": "Personnel costs per employee_2018",
    "Personalkostnader per anställd (tkr)_2019": "Personnel costs per employee_2019",
    "Personalkostnader per anställd (tkr)_2020": "Personnel costs per employee_2020",
    "Personalkostnader per anställd (tkr)_2021": "Personnel costs per employee_2021",
    "Personalkostnader per anställd (tkr)_2022": "Personnel costs per employee_2022",
    "Rörelseresultat (EBITDA)_2012": "EBITDA_2012",
    "Rörelseresultat (EBITDA)_2013": "EBITDA_2013",
    "Rörelseresultat (EBITDA)_2014": "EBITDA_2014",
    "Rörelseresultat (EBITDA)_2015": "EBITDA_2015",
    "Rörelseresultat (EBITDA)_2016": "EBITDA_2016",
    "Rörelseresultat (EBITDA)_2017": "EBITDA_2017",
    "Rörelseresultat (EBITDA)_2018": "EBITDA_2018",
    "Rörelseresultat (EBITDA)_2019": "EBITDA_2019",
    "Rörelseresultat (EBITDA)_2020": "EBITDA_2020",
    "Rörelseresultat (EBITDA)_2021": "EBITDA_2021",
    "Rörelseresultat (EBITDA)_2022": "EBITDA_2022",
    "Nettoomsättningsförändring_2012": "Net revenue change_2012",
    "Nettoomsättningsförändring_2013": "Net revenue change_2013",
    "Nettoomsättningsförändring_2014": "Net revenue change_2014",
    "Nettoomsättningsförändring_2015": "Net revenue change_2015",
    "Nettoomsättningsförändring_2016": "Net revenue change_2016",
    "Nettoomsättningsförändring_2017": "Net revenue change_2017",
    "Nettoomsättningsförändring_2018": "Net revenue change_2018",
    "Nettoomsättningsförändring_2019": "Net revenue change_2019",
    "Nettoomsättningsförändring_2020": "Net revenue change_2020",
    "Nettoomsättningsförändring_2021": "Net revenue change_2021",
    "Nettoomsättningsförändring_2022": "Net revenue change_2022",
    "Du Pont-modellen_2012": "DuPont model_2012",
    "Du Pont-modellen_2013": "DuPont model_2013",
    "Du Pont-modellen_2014": "DuPont model_2014",
    "Du Pont-modellen_2015": "DuPont model_2015",
    "Du Pont-modellen_2016": "DuPont model_2016",
    "Du Pont-modellen_2017": "DuPont model_2017",
    "Du Pont-modellen_2018": "DuPont model_2018",
    "Du Pont-modellen_2019": "DuPont model_2019",
    "Du Pont-modellen_2020": "DuPont model_2020",
    "Du Pont-modellen_2021": "DuPont model_2021",
    "Du Pont-modellen_2022": "DuPont model_2022",
    "Vinstmarginal_2012": "Profit margin_2012",
    "Vinstmarginal_2013": "Profit margin_2013",
    "Vinstmarginal_2014": "Profit margin_2014",
    "Vinstmarginal_2015": "Profit margin_2015",
    "Vinstmarginal_2016": "Profit margin_2016",
    "Vinstmarginal_2017": "Profit margin_2017",
    "Vinstmarginal_2018": "Profit margin_2018",
    "Vinstmarginal_2019": "Profit margin_2019",
    "Vinstmarginal_2020": "Profit margin_2020",
    "Vinstmarginal_2021": "Profit margin_2021",
    "Vinstmarginal_2022": "Profit margin_2022",
    "Bruttovinstmarginal_2012": "Gross profit margin_2012",
    "Bruttovinstmarginal_2013": "Gross profit margin_2013",
    "Bruttovinstmarginal_2014": "Gross profit margin_2014",
    "Bruttovinstmarginal_2015": "Gross profit margin_2015",
    "Bruttovinstmarginal_2016": "Gross profit margin_2016",
    "Bruttovinstmarginal_2017": "Gross profit margin_2017",
    "Bruttovinstmarginal_2018": "Gross profit margin_2018",
    "Bruttovinstmarginal_2019": "Gross profit margin_2019",
    "Bruttovinstmarginal_2020": "Gross profit margin_2020",
    "Bruttovinstmarginal_2021": "Gross profit margin_2021",
    "Bruttovinstmarginal_2022": "Gross profit margin_2022",
    "Rörelsekapital/omsättning_2012": "Working capital/revenue_2012",
    "Rörelsekapital/omsättning_2013": "Working capital/revenue_2013",
    "Rörelsekapital/omsättning_2014": "Working capital/revenue_2014",
    "Rörelsekapital/omsättning_2015": "Working capital/revenue_2015",
    "Rörelsekapital/omsättning_2016": "Working capital/revenue_2016",
    "Rörelsekapital/omsättning_2017": "Working capital/revenue_2017",
    "Rörelsekapital/omsättning_2018": "Working capital/revenue_2018",
    "Rörelsekapital/omsättning_2019": "Working capital/revenue_2019",
    "Rörelsekapital/omsättning_2020": "Working capital/revenue_2020",
    "Rörelsekapital/omsättning_2021": "Working capital/revenue_2021",
    "Rörelsekapital/omsättning_2022": "Working capital/revenue_2022",
    "Soliditet_2012": "Solvency_2012", # Solvency is already given by the data source but it could be calculated as following:
    "Soliditet_2013": "Solvency_2013", # round(100 * (df['Equity_20##'] + 0.786 * df['Untaxed reserves_20##']) / df['Liabilities and equity_20##'], 2)
    "Soliditet_2014": "Solvency_2014",
    "Soliditet_2015": "Solvency_2015",
    "Soliditet_2016": "Solvency_2016",
    "Soliditet_2017": "Solvency_2017",
    "Soliditet_2018": "Solvency_2018",
    "Soliditet_2019": "Solvency_2019",
    "Soliditet_2020": "Solvency_2020",
    "Soliditet_2021": "Solvency_2021",
    "Soliditet_2022": "Solvency_2022",
    "Kassalikviditet_2012": "Quick ratio_2012",
    "Kassalikviditet_2013": "Quick ratio_2013",
    "Kassalikviditet_2014": "Quick ratio_2014",
    "Kassalikviditet_2015": "Quick ratio_2015",
    "Kassalikviditet_2016": "Quick ratio_2016",
    "Kassalikviditet_2017": "Quick ratio_2017",
    "Kassalikviditet_2018": "Quick ratio_2018",
    "Kassalikviditet_2019": "Quick ratio_2019",
    "Kassalikviditet_2020": "Quick ratio_2020",
    "Kassalikviditet_2021": "Quick ratio_2021",
    "Kassalikviditet_2022": "Quick ratio_2022",
    "Anläggningstillgångar_2012": "Fixed assets_2012",
    "Anläggningstillgångar_2013": "Fixed assets_2013",
    "Anläggningstillgångar_2014": "Fixed assets_2014",
    "Anläggningstillgångar_2015": "Fixed assets_2015",
    "Anläggningstillgångar_2016": "Fixed assets_2016",
    "Anläggningstillgångar_2017": "Fixed assets_2017",
    "Anläggningstillgångar_2018": "Fixed assets_2018",
    "Anläggningstillgångar_2019": "Fixed assets_2019",
    "Anläggningstillgångar_2020": "Fixed assets_2020",
    "Anläggningstillgångar_2021": "Fixed assets_2021",
    "Anläggningstillgångar_2022": "Fixed assets_2022",
    "Årets resultat_2012": "Annual result_2012",
    "Årets resultat_2013": "Annual result_2013",
    "Årets resultat_2014": "Annual result_2014",
    "Årets resultat_2015": "Annual result_2015",
    "Årets resultat_2016": "Annual result_2016",
    "Årets resultat_2017": "Annual result_2017",
    "Årets resultat_2018": "Annual result_2018",
    "Årets resultat_2019": "Annual result_2019",
    "Årets resultat_2020": "Annual result_2020",
    "Årets resultat_2021": "Annual result_2021",
    "Årets resultat_2022": "Annual result_2022",
    "Avsättningar (tkr)_2012": "Provisions_2012",
    "Avsättningar (tkr)_2013": "Provisions_2013",
    "Avsättningar (tkr)_2014": "Provisions_2014",
    "Avsättningar (tkr)_2015": "Provisions_2015",
    "Avsättningar (tkr)_2016": "Provisions_2016",
    "Avsättningar (tkr)_2017": "Provisions_2017",
    "Avsättningar (tkr)_2018": "Provisions_2018",
    "Avsättningar (tkr)_2019": "Provisions_2019",
    "Avsättningar (tkr)_2020": "Provisions_2020",
    "Avsättningar (tkr)_2021": "Provisions_2021",
    "Avsättningar (tkr)_2022": "Provisions_2022",
    "Eget kapital_2012": "Equity_2012",
    "Eget kapital_2013": "Equity_2013",
    "Eget kapital_2014": "Equity_2014",
    "Eget kapital_2015": "Equity_2015",
    "Eget kapital_2016": "Equity_2016",
    "Eget kapital_2017": "Equity_2017",
    "Eget kapital_2018": "Equity_2018",
    "Eget kapital_2019": "Equity_2019",
    "Eget kapital_2020": "Equity_2020",
    "Eget kapital_2021": "Equity_2021",
    "Eget kapital_2022": "Equity_2022",
    "Kortfristiga skulder_2012": "Short-term liabilities_2012",
    "Kortfristiga skulder_2013": "Short-term liabilities_2013",
    "Kortfristiga skulder_2014": "Short-term liabilities_2014",
    "Kortfristiga skulder_2015": "Short-term liabilities_2015",
    "Kortfristiga skulder_2016": "Short-term liabilities_2016",
    "Kortfristiga skulder_2017": "Short-term liabilities_2017",
    "Kortfristiga skulder_2018": "Short-term liabilities_2018",
    "Kortfristiga skulder_2019": "Short-term liabilities_2019",
    "Kortfristiga skulder_2020": "Short-term liabilities_2020",
    "Kortfristiga skulder_2021": "Short-term liabilities_2021",
    "Kortfristiga skulder_2022": "Short-term liabilities_2022",
    "Långfristiga skulder_2012": "Long-term liabilities_2012",
    "Långfristiga skulder_2013": "Long-term liabilities_2013",
    "Långfristiga skulder_2014": "Long-term liabilities_2014",
    "Långfristiga skulder_2015": "Long-term liabilities_2015",
    "Långfristiga skulder_2016": "Long-term liabilities_2016",
    "Långfristiga skulder_2017": "Long-term liabilities_2017",
    "Långfristiga skulder_2018": "Long-term liabilities_2018",
    "Långfristiga skulder_2019": "Long-term liabilities_2019",
    "Långfristiga skulder_2020": "Long-term liabilities_2020",
    "Långfristiga skulder_2021": "Long-term liabilities_2021",
    "Långfristiga skulder_2022": "Long-term liabilities_2022",
    "Löner till övriga anställda_2012": "Salaries for other employees_2012",
    "Löner till övriga anställda_2013": "Salaries for other employees_2013",
    "Löner till övriga anställda_2014": "Salaries for other employees_2014",
    "Löner till övriga anställda_2015": "Salaries for other employees_2015",
    "Löner till övriga anställda_2016": "Salaries for other employees_2016",
    "Löner till övriga anställda_2017": "Salaries for other employees_2017",
    "Löner till övriga anställda_2018": "Salaries for other employees_2018",
    "Löner till övriga anställda_2019": "Salaries for other employees_2019",
    "Löner till övriga anställda_2020": "Salaries for other employees_2020",
    "Löner till övriga anställda_2021": "Salaries for other employees_2021",
    "Löner till övriga anställda_2022": "Salaries for other employees_2022",
    "Varav resultatlön till övriga anställda_2012": "Of which bonus salary to other employees_2012",
    "Varav resultatlön till övriga anställda_2013": "Of which bonus salary to other employees_2013",
    "Varav resultatlön till övriga anställda_2014": "Of which bonus salary to other employees_2014",
    "Varav resultatlön till övriga anställda_2015": "Of which bonus salary to other employees_2015",
    "Varav resultatlön till övriga anställda_2016": "Of which bonus salary to other employees_2016",
    "Varav resultatlön till övriga anställda_2017": "Of which bonus salary to other employees_2017",
    "Varav resultatlön till övriga anställda_2018": "Of which bonus salary to other employees_2018",
    "Varav resultatlön till övriga anställda_2019": "Of which bonus salary to other employees_2019",
    "Varav resultatlön till övriga anställda_2020": "Of which bonus salary to other employees_2020",
    "Varav resultatlön till övriga anställda_2021": "Of which bonus salary to other employees_2021",
    "Varav resultatlön till övriga anställda_2022": "Of which bonus salary to other employees_2022",
    "Löner till styrelse & VD_2012": "Salaries for board & CEO_2012",
    "Löner till styrelse & VD_2013": "Salaries for board & CEO_2013",
    "Löner till styrelse & VD_2014": "Salaries for board & CEO_2014",
    "Löner till styrelse & VD_2015": "Salaries for board & CEO_2015",
    "Löner till styrelse & VD_2016": "Salaries for board & CEO_2016",
    "Löner till styrelse & VD_2017": "Salaries for board & CEO_2017",
    "Löner till styrelse & VD_2018": "Salaries for board & CEO_2018",
    "Löner till styrelse & VD_2019": "Salaries for board & CEO_2019",
    "Löner till styrelse & VD_2020": "Salaries for board & CEO_2020",
    "Löner till styrelse & VD_2021": "Salaries for board & CEO_2021",
    "Löner till styrelse & VD_2022": "Salaries for board & CEO_2022",
    "Varav tantiem till styrelse & VD_2012": "Of which bonus to board & CEO_2012",
    "Varav tantiem till styrelse & VD_2013": "Of which bonus to board & CEO_2013",
    "Varav tantiem till styrelse & VD_2014": "Of which bonus to board & CEO_2014",
    "Varav tantiem till styrelse & VD_2015": "Of which bonus to board & CEO_2015",
    "Varav tantiem till styrelse & VD_2016": "Of which bonus to board & CEO_2016",
    "Varav tantiem till styrelse & VD_2017": "Of which bonus to board & CEO_2017",
    "Varav tantiem till styrelse & VD_2018": "Of which bonus to board & CEO_2018",
    "Varav tantiem till styrelse & VD_2019": "Of which bonus to board & CEO_2019",
    "Varav tantiem till styrelse & VD_2020": "Of which bonus to board & CEO_2020",
    "Varav tantiem till styrelse & VD_2021": "Of which bonus to board & CEO_2021",
    "Varav tantiem till styrelse & VD_2022": "Of which bonus to board & CEO_2022",
    "Minoritetsintressen_2015": "Minority interests_2015",
    "Minoritetsintressen_2016": "Minority interests_2016",
    "Minoritetsintressen_2017": "Minority interests_2017",
    "Minoritetsintressen_2018": "Minority interests_2018",
    "Minoritetsintressen_2020": "Minority interests_2020",
    "Nettoomsättning_2012": "Net revenue_2012",
    "Nettoomsättning_2013": "Net revenue_2013",
    "Nettoomsättning_2014": "Net revenue_2014",
    "Nettoomsättning_2015": "Net revenue_2015",
    "Nettoomsättning_2016": "Net revenue_2016",
    "Nettoomsättning_2017": "Net revenue_2017",
    "Nettoomsättning_2018": "Net revenue_2018",
    "Nettoomsättning_2019": "Net revenue_2019",
    "Nettoomsättning_2020": "Net revenue_2020",
    "Nettoomsättning_2021": "Net revenue_2021",
    "Nettoomsättning_2022": "Net revenue_2022",
    "Obeskattade reserver_2012": "Untaxed reserves_2012",
    "Obeskattade reserver_2013": "Untaxed reserves_2013",
    "Obeskattade reserver_2014": "Untaxed reserves_2014",
    "Obeskattade reserver_2015": "Untaxed reserves_2015",
    "Obeskattade reserver_2016": "Untaxed reserves_2016",
    "Obeskattade reserver_2017": "Untaxed reserves_2017",
    "Obeskattade reserver_2018": "Untaxed reserves_2018",
    "Obeskattade reserver_2019": "Untaxed reserves_2019",
    "Obeskattade reserver_2020": "Untaxed reserves_2020",
    "Obeskattade reserver_2021": "Untaxed reserves_2021",
    "Obeskattade reserver_2022": "Untaxed reserves_2022",
    "Omsättning_2012": "Revenue_2012",
    "Omsättning_2013": "Revenue_2013",
    "Omsättning_2014": "Revenue_2014",
    "Omsättning_2015": "Revenue_2015",
    "Omsättning_2016": "Revenue_2016",
    "Omsättning_2017": "Revenue_2017",
    "Omsättning_2018": "Revenue_2018",
    "Omsättning_2019": "Revenue_2019",
    "Omsättning_2020": "Revenue_2020",
    "Omsättning_2021": "Revenue_2021",
    "Omsättning_2022": "Revenue_2022",
    "Omsättningstillgångar_2012": "Current assets_2012",
    "Omsättningstillgångar_2013": "Current assets_2013",
    "Omsättningstillgångar_2014": "Current assets_2014",
    "Omsättningstillgångar_2015": "Current assets_2015",
    "Omsättningstillgångar_2016": "Current assets_2016",
    "Omsättningstillgångar_2017": "Current assets_2017",
    "Omsättningstillgångar_2018": "Current assets_2018",
    "Omsättningstillgångar_2019": "Current assets_2019",
    "Omsättningstillgångar_2020": "Current assets_2020",
    "Omsättningstillgångar_2021": "Current assets_2021",
    "Omsättningstillgångar_2022": "Current assets_2022",
    "Övrig omsättning_2012": "Other revenue_2012",
    "Övrig omsättning_2013": "Other revenue_2013",
    "Övrig omsättning_2014": "Other revenue_2014",
    "Övrig omsättning_2015": "Other revenue_2015",
    "Övrig omsättning_2016": "Other revenue_2016",
    "Övrig omsättning_2017": "Other revenue_2017",
    "Övrig omsättning_2018": "Other revenue_2018",
    "Övrig omsättning_2019": "Other revenue_2019",
    "Övrig omsättning_2020": "Other revenue_2020",
    "Övrig omsättning_2021": "Other revenue_2021",
    "Övrig omsättning_2022": "Other revenue_2022",
    "Resultat efter finansnetto_2012": "Profit after financial items_2012",
    "Resultat efter finansnetto_2013": "Profit after financial items_2013",
    "Resultat efter finansnetto_2014": "Profit after financial items_2014",
    "Resultat efter finansnetto_2015": "Profit after financial items_2015",
    "Resultat efter finansnetto_2016": "Profit after financial items_2016",
    "Resultat efter finansnetto_2017": "Profit after financial items_2017",
    "Resultat efter finansnetto_2018": "Profit after financial items_2018",
    "Resultat efter finansnetto_2019": "Profit after financial items_2019",
    "Resultat efter finansnetto_2020": "Profit after financial items_2020",
    "Resultat efter finansnetto_2021": "Profit after financial items_2021",
    "Resultat efter finansnetto_2022": "Profit after financial items_2022",
    "Rörelseresultat (EBIT)_2012": "EBIT_2012",
    "Rörelseresultat (EBIT)_2013": "EBIT_2013",
    "Rörelseresultat (EBIT)_2014": "EBIT_2014",
    "Rörelseresultat (EBIT)_2015": "EBIT_2015",
    "Rörelseresultat (EBIT)_2016": "EBIT_2016",
    "Rörelseresultat (EBIT)_2017": "EBIT_2017",
    "Rörelseresultat (EBIT)_2018": "EBIT_2018",
    "Rörelseresultat (EBIT)_2019": "EBIT_2019",
    "Rörelseresultat (EBIT)_2020": "EBIT_2020",
    "Rörelseresultat (EBIT)_2021": "EBIT_2021",
    "Rörelseresultat (EBIT)_2022": "EBIT_2022",
    "Skulder och eget kapital_2012": "Liabilities and equity_2012",
    "Skulder och eget kapital_2013": "Liabilities and equity_2013",
    "Skulder och eget kapital_2014": "Liabilities and equity_2014",
    "Skulder och eget kapital_2015": "Liabilities and equity_2015",
    "Skulder och eget kapital_2016": "Liabilities and equity_2016",
    "Skulder och eget kapital_2017": "Liabilities and equity_2017",
    "Skulder och eget kapital_2018": "Liabilities and equity_2018",
    "Skulder och eget kapital_2019": "Liabilities and equity_2019",
    "Skulder och eget kapital_2020": "Liabilities and equity_2020",
    "Skulder och eget kapital_2021": "Liabilities and equity_2021",
    "Skulder och eget kapital_2022": "Liabilities and equity_2022",
    "Sociala kostnader_2012": "Social costs_2012",
    "Sociala kostnader_2013": "Social costs_2013",
    "Sociala kostnader_2014": "Social costs_2014",
    "Sociala kostnader_2015": "Social costs_2015",
    "Sociala kostnader_2016": "Social costs_2016",
    "Sociala kostnader_2017": "Social costs_2017",
    "Sociala kostnader_2018": "Social costs_2018",
    "Sociala kostnader_2019": "Social costs_2019",
    "Sociala kostnader_2020": "Social costs_2020",
    "Sociala kostnader_2021": "Social costs_2021",
    "Sociala kostnader_2022": "Social costs_2022",
    "Tecknat ej inbetalt kapital_2012": "Subscribed unpaid capital_2012",
    "Tecknat ej inbetalt kapital_2013": "Subscribed unpaid capital_2013",
    "Tecknat ej inbetalt kapital_2014": "Subscribed unpaid capital_2014",
    "Tecknat ej inbetalt kapital_2015": "Subscribed unpaid capital_2015",
    "Tecknat ej inbetalt kapital_2016": "Subscribed unpaid capital_2016",
    "Tecknat ej inbetalt kapital_2017": "Subscribed unpaid capital_2017",
    "Tecknat ej inbetalt kapital_2018": "Subscribed unpaid capital_2018",
    "Tecknat ej inbetalt kapital_2019": "Subscribed unpaid capital_2019",
    "Tecknat ej inbetalt kapital_2020": "Subscribed unpaid capital_2020",
    "Tecknat ej inbetalt kapital_2021": "Subscribed unpaid capital_2021",
    "Tecknat ej inbetalt kapital_2022": "Subscribed unpaid capital_2022",
    "Utdelning till aktieägare_2012": "Dividends_2012",
    "Utdelning till aktieägare_2013": "Dividends_2013",
    "Utdelning till aktieägare_2014": "Dividends_2014",
    "Utdelning till aktieägare_2015": "Dividends_2015",
    "Utdelning till aktieägare_2016": "Dividends_2016",
    "Utdelning till aktieägare_2017": "Dividends_2017",
    "Utdelning till aktieägare_2018": "Dividends_2018",
    "Utdelning till aktieägare_2019": "Dividends_2019",
    "Utdelning till aktieägare_2020": "Dividends_2020",
    "Utdelning till aktieägare_2021": "Dividends_2021",
    "Utdelning till aktieägare_2022": "Dividends_2022"
                                    }

Activity_types_translation_table = {
    'Alkoholhaltiga drycker, Butikshandel': 'Alcoholic Beverages, Retail Trade',
    'Anläggningsarbeten': 'Construction Works',
    'Antikviteter & Beg. Böcker, Butikshandel': 'Antiques & Second-Hand Books, Retail Trade',
    'Apotekshandel': 'Pharmacy Retail Trade',
    'Arbetsförmedling & Rekrytering': 'Employment Agencies & Recruitment',
    'Arkitektverksamhet': 'Architectural Services',
    'Auktioner': 'Auctions',
    'Band & Skivor, Butikshandel': 'Music & Records, Retail Trade',
    'Barnkläder, Butikshandel': 'Children''s Clothing, Retail Trade',
    'Begagnade varor övriga, Butikshandel': 'Used Goods, Other, Retail Trade',
    'Belysningsartiklar, Butikshandel': 'Lighting Articles, Retail Trade',
    'Blommor & Växter, Butikshandel': 'Flowers & Plants, Retail Trade',
    'Bröd & Konditori, Butikshandel': 'Bread & Pastry, Retail Trade',
    'Butikshandel, övrig': 'Retail Trade, Other',
    'Byggnadssnickeriarbeten': 'Building Carpentry Works',
    'Byggverksamhet': 'Construction',
    'Båtar, Butikshandel': 'Boats, Retail Trade',
    'Böcker, Butikshandel': 'Books, Retail Trade',
    'Campingplatsverksamhet': 'Camping Site Operation',
    'Catering': 'Catering',
    'Centralkök för sjukhus': 'Central Kitchen for Hospitals',
    'Centralkök för skolor & omsorg': 'Central Kitchen for Schools & Care',
    'Cyklar, Butikshandel': 'Bicycles, Retail Trade',
    'Damkläder, Butikshandel': 'Women''s Clothing, Retail Trade',
    'Datakonsultverksamhet': 'Data Consulting Services',
    'Dataprogrammering': 'Data Programming',
    'Datordrifttjänster': 'Data Hosting Services',
    'Datorer, Program, Data-& TV-spel, Butikshandel': 'Computers, Software, Data & Video Games, Retail Trade',
    'Detaljhandel, övrig': 'Retail Trade, Other',
    'Drivmedel, Detaljhandel': 'Fuel, Retail Trade',
    'El-VVS & Bygginstallationer': 'Electrical, Plumbing & Building Installations',
    'Elektriska Hushållsmaskiner, Butikshandel': 'Electrical Household Appliances, Retail Trade',
    'Fastighetsförmedling': 'Real Estate Agency',
    'Fastighetsförvaltning på uppdrag': 'Real Estate Management on Behalf',
    'Fastighetsrelaterade stödtjänster': 'Real Estate-Related Support Services',
    'Finans, administrativa tjänster': 'Finance, Administrative Services',
    'Finansförmedling, övrig': 'Financial Intermediation, Other',
    'Finansiell leasing': 'Financial Leasing',
    'Finansiella stödtjänster, övriga': 'Other Financial Support Services',
    'Fisk & Skaldjur, Butikshandel': 'Fish & Seafood, Retail Trade',
    'Fondanknuten livförsäkring': 'Unit-Linked Life Insurance',
    'Fonder & liknande finansiella enheter, övriga': 'Funds & Similar Financial Entities, Other',
    'Fondförvaltning, övrig': 'Fund Management, Other',
    'Fotoutrustning, Butikshandel': 'Photographic Equipment, Retail Trade',
    'Frukt & Grönsaker, Butikshandel': 'Fruits & Vegetables, Retail Trade',
    'Färg & Lack, Butikshandel': 'Paint & Varnish, Retail Trade',
    'Försäkring & Pensionsfond stödtjänster, övriga': 'Insurance & Pension Fund Support Services, Other',
    'Förvaltning & Handel med Värdepapper': 'Management & Trading of Securities',
    'Förvaltning & Handel med värdepapper': 'Securities Management & Trading',
    'Förvaltning av investeringsfonder': 'Investment Fund Management',
    'Förvaltning i Bostadsrättsföreningar': 'Management in Housing Cooperatives',
    'Glas & Porslin, Butikshandel': 'Glass & Porcelain, Retail Trade',
    'Glasmästeriarbeten': 'Glass Cutting Works',
    'Golv- & Väggbeläggningsarbeten': 'Flooring & Wall Covering Works',
    'Grafisk Designverksamhet': 'Graphic Design Services',
    'Guldsmedsvaror & Smycken, Butikshandel': 'Goldsmith''s Goods & Jewelry, Retail Trade',
    'Handel med egna fastigheter': 'Trade with Own Properties',
    'Herrkläder, Butikshandel': 'Men''s Clothing, Retail Trade',
    'Holdingverksamhet i finansiella koncerner': 'Holding Company in Financial Conglomerates',
    'Holdingverksamhet i icke-finansiella koncerner': 'Holding Company in Non-Financial Conglomerates',
    'Hotell & Restaurang': 'Hotel & Restaurant',
    'Hälso- & Sjukvård, övriga': 'Healthcare, Other',
    'Hälsokost, Butikshandel': 'Health Food, Retail Trade',
    'IT- & Datatjänster, övriga': 'IT & Data Services, Other',
    'Industri- & Produktdesignverksamhet': 'Industrial & Product Design Services',
    'Inredningsarkitekt': 'Interior Architect',
    'Inredningstextilier, Butikshandel': 'Interior Textiles, Retail Trade',
    'Investeringsfonder': 'Investment Funds',
    'Investment- & Riskkapitalbolag': 'Investment & Venture Capital Companies',
    'Järn- & VVS- varor, Butikshandel': 'Iron & Plumbing Goods, Retail Trade',
    'Klockor, Butikshandel': 'Watches, Retail Trade',
    'Kläder, Butikshandel': 'Clothing, Retail Trade',
    'Konfektyr, Butikshandel': 'Confectionery, Retail Trade',
    'Konferensanläggningar': 'Conference Facilities',
    'Konst & Galleri, Butikshandel': 'Art & Gallery, Retail Trade',
    'Kontorsförbrukningsvaror, Butikshandel': 'Office Consumables, Retail Trade',
    'Kontorsmöbler, Butikshandel': 'Office Furniture, Retail Trade',
    'Kosmetika & Hygienartiklar, Butikshandel': 'Cosmetics & Hygiene Articles, Retail Trade',
    'Kreditgivning, övrig': 'Credit Provision, Other',
    'Kött & Charkuterier, Butikshandel': 'Meat & Charcuterie, Retail Trade',
    'Livförsäkring, övrig': 'Life Insurance, Other',
    'Livsmedel övriga, Butikshandel': 'Other Foodstuffs, Retail Trade',
    'Livsmedelshandel': 'Food Retail Trade',
    'Ljud & Bild, Butikshandel': 'Audio & Video, Retail Trade',
    'Logiverksamhet, övrig': 'Logistics Services, Other',
    'Markundersökning': 'Market Research',
    'Mattor & Väggbeklädnad, Butikshandel': 'Carpets & Wall Coverings, Retail Trade',
    'Monetär finansförmedling, övrig': 'Monetary Financial Intermediation, Other',
    'Musikinstrument & Noter, Butikshandel': 'Musical Instruments & Scores, Retail Trade',
    'Mynt & Frimärken, Butikshandel': 'Coins & Stamps, Retail Trade',
    'Måleriarbeten': 'Painting Works',
    'Möbler för hemmet, Butikshandel': 'Furniture for the Home, Retail Trade',
    'Optiker, Butikshandel': 'Optician, Retail Trade',
    'Pensionsfondsverksamhet': 'Pension Fund Activities',
    'Personalfunktioner, övrigt': 'Personnel Functions, Other',
    'Personalmatsalar': 'Employee Cafeterias',
    'Personaluthyrning': 'Staff Leasing',
    'Postorder- & Internethandel': 'Mail Order & Internet Trade',
    'Primärvårdsmottagning': 'Primary Care Clinic',
    'Puts-, Fasad- & Stuckatörsarbeten': 'Cleaning & Facility Management',
    'Pälsar, Butikshandel': 'Furs, Retail Trade',
    'Rengöring & Lokalvård': 'Cleaning & Facility Management',
    'Restaurangverksamhet': 'Restaurant Operations',
    'Risk- & Skadebedömning': 'Risk & Damage Assessment',
    'Rivning av hus & byggnader': 'Demolition of Houses & Buildings',
    'Sjukvårdsartiklar, Butikshandel': 'Medical Supplies, Retail Trade',
    'Skadeförsäkring': 'Insurance Claims',
    'Skor, Butikshandel': 'Shoes, Retail Trade',
    'Skötsel & Underhåll av Grönytor': 'Care & Maintenance of Green Areas',
    'Slutbehandling av byggnader': 'Finalization of Buildings',
    'Sluten Sjukvård': 'Closed Healthcare',
    'Sociala insatser': 'Social Services',
    'Spel & Leksaker, Butikshandel': 'Games & Toys, Retail Trade',
    'Sport- & Fritidsartiklar, Butikshandel': 'Sports & Leisure Articles, Retail Trade',
    'Stugbyverksamhet': 'Cottage Rentals',
    'Sällskapsdjur, Butikshandel': 'Companion Animals, Retail Trade',
    'Takarbeten': 'Roofing',
    'Teknisk konsult inom Bygg- & Anläggningsteknik': 'Technical Consultancy in Building & Construction Engineering',
    'Telekommunikation, Satellit': 'Telecommunication, Satellite',
    'Telekommunikation, Trådbunden': 'Telecommunication, Wired',
    'Telekommunikation, Trådlös': 'Telecommunication, Wireless',
    'Telekommunikation, övrig': 'Telecommunication, Other',
    'Telekommunikationsutrustning, Butikshandel': 'Telecommunications Equipment, Retail Trade',
    'Textilier, Butikshandel': 'Textiles, Retail Trade',
    'Tidningar, Butikshandel': 'Newspapers, Retail Trade',
    'Tobaksvaror, Butikshandel': 'Tobacco Products, Retail Trade',
    'Torg- & Marknadshandel': 'Square & Market Trade',
    'Utformning av Byggprojekt': 'Design of Building Projects',
    'Utgivning av annan programvara': 'Publication of Other Software',
    'Utgivning av dataspel': 'Publication of Video Games',
    'Uthyrning & Förvaltning av Fastigheter': 'Rental & Management of Properties',
    'Uthyrning av Bygg- & Anläggningsmaskiner med förare': 'Rental of Building & Construction Machinery with Operator',
    'Vandrarhemsverksamhet': 'Hostel Operations',
    'Varuhus- & Stormarknadshandel': 'Department Store & Hypermarket Trade',
    'Verksamhet utförd av Försäkringsombud & Försäkringsmäklare': 'Activities carried out by Insurance Agents & Brokers',
    'Verksamhet utförd av Värdepappers- & Varumäklare': 'Activities carried out by Securities & Commodities Brokers',
    'Virke & Byggvaror, Butikshandel': 'Timber & Building Materials, Retail Trade',
    'Väskor & Reseffekter, Butikshandel': 'Bags & Travel Items, Retail Trade',
    'Vård & Omsorg med Boende': 'Healthcare & Care with Accommodation',
    'Återförsäkring': 'Reinsurance',
    'Bränsle, Mineraler & Industrikem. Partihandel': 'Fuel, Minerals & Industrial Chemicals Wholesale',
    'Datorer, Program & Kringutr, Partihandel': 'Computers, Software & Peripherals Wholesale',
    'Drivning': 'Driving',
    'Elektronikindustri': 'Electronics Industry',
    'Elektronikkomponenter, Partihandel': 'Electronic Components Wholesale',
    'Elgenerering': 'Electric Generation',
    'Farmaceutiska basprodukter, tillverkning': 'Pharmaceutical Raw Materials Manufacturing',
    'Gruv-, Bygg- & Anläggningsmaskiner, Partihandel': 'Mining, Construction & Construction Machinery Wholesale',
    'Hushållsapparater & Elartiklar, Partihandel': 'Household Appliances & Electrical Articles Wholesale',
    'Insatsvaror övriga, Partihandel': 'Other Input Goods, Wholesale',
    'Järnhandelsvaror, Partihandel': 'Ironmongery Wholesale',
    'Kemiska produkter, tillverkning': 'Chemical Products Manufacturing',
    'Konsultverksamhet avseende företags org.': 'Consultancy on Corporate Organization',
    'Kontors- & Butiksinred, tillverkning': 'Office & Shop Fittings Manufacturing',
    'Ljud, Bild & Videoutrustning, Partihandel': 'Audio, Video & Visual Equipment Wholesale',
    'Läkemedel, tillverkning': 'Pharmaceutical Manufacturing',
    'Metallindustri': 'Metal Industry',
    'Pappers- & Pappersvarutillverkning': 'Paper & Paper Product Manufacturing',
    'Partihandel, övrig': 'Wholesale Trade, Other',
    'Personbilar & Lätta Motorfordon, Handel': 'Cars & Light Motor Vehicles, Retail Trade',
    'Rälsfordon, tillverkning': 'Rail Vehicles Manufacturing',
    'Skogsförvaltning': 'Forest Management',
    'Specialistläkare ej på sjukhus': 'Specialist Doctors not in Hospitals',
    'Specialistläkare på sjukhus': 'Specialist Doctors in Hospitals',
    'Spel- & Vadhållningsverksamhet': 'Gambling & Betting',
    'Stats- & Kommunledning, Lagstiftning & Planering': 'State & Municipal Management, Legislation & Planning',
    'Sågning, Hyvling & Impregnering': 'Sawing, Planing & Impregnation',
    'Tandläkare': 'Dentists',
    'Transport stödtjänster, övriga': 'Transport Support Services, Other',
    'Transportmedelsindustri': 'Transport Industry',
    'Verksamheter som utövas av huvudkontor': 'Activities of Head Offices',
    'Veterinärverksamhet': 'Veterinary Activities',
    'Virke & Byggmaterial, Partihandel': 'Timber & Building Materials, Wholesale',
    'Värme & Kyla': 'Heat & Cold',
    'Öppen Hälso- & Sjukvård': 'Open Health & Medical Care',
    'Kemiska produkter, Partihandel': 'Chemical Products, Wholesale',
    'Juridik, Ekonomi, Vetenskap & Teknik, övrig': 'Law, Economics',
    'nan': 'Nanotechnology',
    'Vägtransport, Godstrafik': 'Road Transport, Freight Traffic',
    'Verktygsmaskiner, tillverkning': 'Machine Tools Manufacturing',
    'Vatten & Avlopp': 'Water & Sewerage',
    'VVS-varor, Partihandel': 'HVAC Goods, Wholesale',
    'Uthyrning & Leasing av Bygg- & Anläggningsmaskiner': 'Rental & Leasing of Building & Construction Machinery',
    'Trävaror, tillverkning': 'Wood Products Manufacturing',
    'Trähus, tillverkning': 'Wooden Houses Manufacturing',
    'Teknisk konsult inom Industriteknik': 'Technical Consultancy in Industrial Engineering',
    'Teknisk konsult inom Elteknik': 'Technical Consultancy in Electrical Engineering',
    'Taxi': 'Taxi',
    'Plastförpackningstillverkning': 'Plastic Packaging Manufacturing',
    'Möbler, Mattor & Belysning, Partihandel': 'Furniture, Carpets & Lighting',
    'Motorcyklar, Handel, service & tillbehör': 'Motorcycles, Trade',
    'Maskiner, tillverkning': 'Machinery Manufacturing',
    'Maskiner, reparation': 'Machinery Repair',
    'Maskiner & Utrustning övriga, Partihandel': 'Other Machinery & Equipment, Wholesale',
    'Lastbilar, Bussar & Specialfordon, Handel': 'Trucks, Buses & Special Vehicles',
    'Köttprodukter': 'Meat Products',
    'Kläder & Skor, Partihandel': 'Clothing & Shoes, Wholesale',
    'Juridisk verksamhet, övrig': 'Legal Activities, Other',
    'Jord- & Skogsbruksmaskiner, tillverkning': 'Agricultural & Forestry Machinery Manufacturing',
    'Husvagnar, Husbilar & Släp, Handel': 'Caravans, Motorhomes & Trailers, Retail Trade',
    'Gummivaror, tillverkning': 'Rubber Goods Manufacturing',
    'Glasstillverkning': 'Glass Manufacturing',
    'Elektriska Komponenter & Kretskort, tillverkning': 'Electrical Components & Circuit Boards Manufacturing',
    'Elartiklar, Partihandel': 'Electrical Articles, Wholesale',
    'Elapparatur, Reparation': 'Electrical Equipment, Repair',
    'Byggnadsmetallvaror, tillverkning': 'Building Metal Goods Manufacturing',
    'Yrkesförarutbildning': 'Professional Driver Training',
    'Verktyg & Redskap, tillverkning': 'Tools & Implements Manufacturing',
    'Tryckning av Böcker & Övrigt': 'Book Printing & Other',
    'Teknisk konsult inom Energi-, Miljö- & VVS-teknik': 'Technical Consultancy in Energy, Environmental & HVAC Engineering',
    'Teknisk Provning & Analys': 'Technical Testing & Analysis',
    'Säkerhetssystemtjänster': 'Security System Services',
    'Stenhuggning': 'Stone Cutting',
    'Skogsskötsel': 'Forestry Management',
    'Skogsbruk': 'Forestry',
    'Sjötransport, stödtjänster': 'Maritime Transport Support Services',
    'Reklam, PR, Mediebyrå & Annonsförsälj.': 'Advertising, PR, Media Agency & Ad Sales',
    'Redovisning & bokföring': 'Accounting & Bookkeeping',
    'Plastvarutillverkning, övrig': 'Plastic Goods Manufacturing, Other',
    'Plastvaror, tillverkning': 'Plastic Goods Manufacturing',
    'Personalutbildning': 'Staff Training',
    'Mät- & Precisionsinstrument, Partihandel': 'Measurement & Precision Instruments, Wholesale',
    'Motorfordon, reparation & underhåll': 'Motor Vehicles, Repair & Maintenance',
    'Mjölkproduktion & Nötkreatursuppfödning': 'Milk Production & Cattle Farming',
    'Mineralutvinning, övrig': 'Mineral Extraction, Other',
    'Marknads- & Opinionsundersökning': 'Market & Opinion Research',
    'Livsmedelsframställning': 'Food Production',
    'Landtransport, stödtjänster': 'Land Transport Support Services',
    'Landtransport av passagerare, övrig': 'Passenger Land Transport, Other',
    'Källsorterat material': 'Source-Sorted Material',
    'Kroppsvård': 'Personal Care',
    'Jordbruksmaskiner, Partihandel': 'Agricultural Machinery, Wholesale',
    'Industri- Maskiner & Utrustning, installation': 'Industrial Machinery & Equipment, Installation',
    'Frukt & Grönsaker, Partihandel': 'Fruits & Vegetables, Wholesale',
    'Flyttjänster': 'Moving Services',
    'Fiskodling': 'Fish Farming',
    'Fisk, Skalddjur & andra livsmedel, Partihandel': 'Fish, Shellfish & Other Food, Wholesale',
    'Dörrar av Trä, tillverkning': 'Wooden Doors Manufacturing',
    'Databehandling & Hosting': 'Data Processing & Hosting',
    'Båt & Fartyg, tillverkning': 'Boat & Ship Manufacturing',
    'Byggplastvarutillverkning': 'Building Plastic Goods Manufacturing',
    'Webbportaler': 'Web Portals',
    'Verktygsmaskiner, Partihandel': 'Machine Tools, Wholesale',
    'Utbildning, övrig': 'Education, Other',
    'Transportmedel övriga, tillverkning': 'Other Transport Equipment, Manufacturing',
    'Textilier, Kläder & Skodon, Partihandel': 'Textiles, Clothing & Footwear, Wholesale',
    'Teknisk konsultverksamhet, övrig': 'Technical Consultancy, Other',
    'Sportverksamhet, övrig': 'Sports Activities, Other',
    'Sport- & Fritidsartiklar, Partihandel': 'Sports & Leisure Articles, Wholesale',
    'Specialsortiment': 'Special Assortment',
    'Smyckestillverkning': 'Jewelry Manufacturing',
    'Skidsportanläggningar': 'Ski Resorts',
    'Service till husdjursskötsel': 'Pet Care Services',
    'Service till Skogsbruk': 'Forestry Services',
    'Plantskoleväxter, odling': 'Nursery Plants, Cultivation',
    'Möbler övriga, tillverkning': 'Other Furniture Manufacturing',
    'Motorfordon, reservdelar & tillbehör': 'Motor Vehicles, Parts & Accessories',
    'Medicinsk utrustning & Apoteksvaror, Partihandel': 'Medical Equipment & Pharmacy Goods, Wholesale',
    'Litterärt & Konstnärligt skapande': 'Literary & Artistic Creation',
    'Kött & Köttvaror, Partihandel': 'Meat & Meat Products, Wholesale',
    'Köksmöbler, tillverkning': 'Kitchen Furniture Manufacturing',
    'Kultur, Nöje & Fritid': 'Culture, Entertainment & Leisure',
    'Kontorsförbrukningsvaror, Partihandel': 'Office Supplies, Wholesale',
    'Kläder & Textilier, tillverkning': 'Clothing & Textiles, Manufacturing',
    'Järnmalmsutvinning': 'Iron Ore Extraction',
    'Industriförnödenheter, Partihandel': 'Industrial Supplies, Wholesale',
    'Idrottsplatser & Sportanläggningar': 'Sports Facilities & Sports Centers',
    'Hushållsvaror övriga, Partihandel': 'Household Goods, Other, Wholesale',
    'Hushålls- & Personartiklar, reparation, övriga': 'Household & Personal Items, Repair, Other',
    'Frisörer': 'Hairdressers',
    'Callcenterverksamhet': 'Call Center Operations',
    'Bioteknisk Forskning & Utveckling': 'Biotechnical Research & Development',
    'Betongvarutillverkning': 'Concrete Goods Manufacturing',
    'Avfallshantering & Återvinning': 'Waste Management & Recycling',
    'Artistisk verksamhet': 'Artistic Activities',
    'Advokatbyråer': 'Law Firms',
    'Översättning & Tolkning': 'Translation & Interpretation',
    'Äggproduktion': 'Egg Production',
    'Vård av historiska Minnesmärken & Byggnader': 'Care of Historical Monuments & Buildings',
    'Växtodling': 'Crop Farming',
    'Växter, bearbetning': 'Plants, Processing',
    'Vävnadstillverkning': 'Tissue Manufacturing',
    'Virkesmätning': 'Timber Measurement',
    'Vapen & Ammunition, tillverkning': 'Weapons & Ammunition, Manufacturing',
    'Utvinning, stödtjänster': 'Extraction, Support Services',
    'Uttjänta fordon, Partihandel': 'Used Vehicles, Wholesale',
    'Utrustning Reparation, övrig': 'Equipment Repair, Other',
    'Utrikesförvaltning': 'Foreign Affairs',
    'Uthyrning av Videokassetter & Dvd-skivor': 'Rental of Videotapes & DVD Discs',
    'Uthyrning & Leasing övrigt': 'Rental & Leasing, Other',
    'Uthyrning & Leasing av flygplan': 'Rental & Leasing of Aircraft',
    'Uthyrning & Leasing av andra Hushållsartiklar & Varor för Personligt bruk': 'Rental & Leasing of Other Household Items & Personal Use Goods',
    'Uthyrning & Leasing av Personbilar & lätta Motorfordon': 'Rental & Leasing of Cars & Light Motor Vehicles',
    'Uthyrning & Leasing av Lastbilar & andra tunga Motorfordon': 'Rental & Leasing of Trucks & Other Heavy Motor Vehicles',
    'Uthyrning & Leasing av Kontorsmaskiner & Kontorsutrustning (inklusive datorer)': 'Rental & Leasing of Office Machinery & Equipment (including computers)',
    'Uthyrning & Leasing av Jordbruksmaskiner & Jordbruksredskap': 'Rental & Leasing of Agricultural Machinery & Equipment',
    'Uthyrning & Leasing av Fritids- & Sportutrustning': 'Rental & Leasing of Recreational & Sports Equipment',
    'Uthyrning & Leasing av Fartyg & Båtar': 'Rental & Leasing of Ships & Boats',
    'Utgivning av tidskrifter': 'Magazine Publishing',
    'Utbildningsväsendet, stödverksamhet': 'Education, Support Activities',
    'Urtillverkning': 'Clock Manufacturing',
    'Uran- & Toriummalmutvinning': 'Uranium & Thorium Ore Mining',
    'Ur & Guldssmedsvaror, Partihandel': 'Clocks & Goldsmith Goods, Wholesale',
    'Ur & Guldsmedsvaror, reparation': 'Clocks & Goldsmith Goods, Repair',
    'Universitets- & Högskoleutbildning samt Forskning': 'University & Higher Education and Research',
    'Universitets- & Högskoleutbildning': 'University & Higher Education',
    'Tävling med hästar': 'Horse Racing',
    'Tvål, Såpa & Tvättmedel, tillverkning': 'Soap & Detergent Manufacturing',
    'Tvätteriverksamhet': 'Laundry Operations',
    'Turist- & Bokningsservice': 'Tourist & Booking Services',
    'Trädgårdar, Djurparker & Naturreservat, drift': 'Gardens, Zoos & Nature Reserves, Operation',
    'Tryckning av Tidsskrifter': 'Printing of Journals',
    'Tryckning av Dagstidningar': 'Printing of Newspapers',
    'Trav- & Galoppbanor': 'Trotting & Galloping Tracks',
    'Transportmedel övriga, reparation': 'Other Transport Equipment, Repair',
    'Trafikskoleverksamhet': 'Driving School Operations',
    'Torvutvinning': 'Peat Extraction',
    'Tobaksvarutillverkning': 'Tobacco Product Manufacturing',
    'Tobaksodling': 'Tobacco Cultivation',
    'Tobak, Partihandel': 'Tobacco, Wholesale',
    'Tillverkning, övrig': 'Manufacturing, Other',
    'Textilier, Partihandel': 'Textiles, Wholesale',
    'Textil-, Sy- & Stickmaskiner, Partihandel': 'Textile, Sewing & Knitting Machines, Wholesale',
    'Teleprodukter, Partihandel': 'Telecom Products, Wholesale',
    'Te- & Kaffetillverkning': 'Tea & Coffee Production',
    'Tandprotestillverkning': 'Denture Production',
    'TV-program planering': 'TV Program Planning',
    'Sötvattensfiske': 'Freshwater Fishing',
    'Sällskapsdjur, uppfödning': 'Companion Animal Breeding',
    'Säkerhetsverksamhet': 'Security Operations',
    'Syntetiskt basgummi, tillverkning': 'Synthetic Base Rubber Manufacturing',
    'Svampodling': 'Mushroom Cultivation',
    'Stärkelsetillverkning': 'Starch Production',
    'Studieförbundens & Frivilligorg. utbildning': 'Study Associations & Voluntary Organization Education',
    'Strumpor, tillverkning': 'Socks Manufacturing',
    'Sprängämne, tillverkning': 'Explosive Manufacturing',
    'Sportklubbars & Idrottsför.  verksamhet': 'Sports Clubs & Athletic Association Activities',
    'Sportartikelstillverkning': 'Sports Article Manufacturing',
    'Sport- & Fritidsutbildning': 'Sports & Leisure Education',
    'Spel- & Leksakstillverkning': 'Game & Toy Manufacturing',
    'Spannmål, Balj- & Oljeväxter, odling': 'Cereal, Legume & Oilseed Cultivation',
    'Socketbetsodling': 'Safflower Cultivation',
    'Sockertillverkning': 'Sugar Manufacturing',
    'Socker, Choklad & Sockerkonfekt, Partihandel': 'Sugar, Chocolate & Confectionery, Wholesale',
    'Smågrisuppfödning': 'Piglet Farming',
    'Slaktsvinuppfödning': 'Pork Production',
    'Skönhetsvård': 'Beauty Care',
    'Skor, tillverkning': 'Shoe Manufacturing',
    'Skomakare': 'Cobbler',
    'Skatterådgivning': 'Tax Consultancy',
    'Service till växtodling': 'Crop Service',
    'Sanitetsgods, Partihandel': 'Sanitary Goods, Wholesale',
    'Samhällsvet. & Humanistisk F&U': 'Social Sciences & Humanities R&D',
    'Samhällelig informationsförsörjning': 'Social Information Supply',
    'Saltvattensfiske': 'Saltwater Fishing',
    'Råpetroleumutvinning': 'Crude Petroleum Extraction',
    'Revision': 'Audit',
    'Resebyråer': 'Travel Agencies',
    'Researrangemang': 'Travel Arrangements',
    'Reproduktion av inspelningar': 'Reproduction of Recordings',
    'Renskötsel': 'Reindeer Husbandry',
    'Religiösa samfund': 'Religious Communities',
    'Reklamfotoverksamhet': 'Advertising Photography',
    'Radiosändning': 'Radio Broadcasting',
    'Pälsvaror, tillverkning': 'Fur Goods Manufacturing',
    'Publicering av Kataloger & Sändlistor': 'Publication of Catalogs & Broadcasting Schedules',
    'Press- & Övrig Fotografverksamhet': 'Press & Other Photography Activities',
    'Potatisodling': 'Potato Cultivation',
    'Potatisberedning': 'Potato Processing',
    'Porträttfotoverksamhet': 'Portrait Photography',
    'Planglas': 'Flat Glass',
    'Petroleumraffinering': 'Petroleum Refining',
    'Personalförvaltning & andra stödtjänster': 'Personnel Management & Other Support Services',
    'Patentbyråer': 'Patent Agencies',
    'Partihandel': 'Wholesale Trade',
    'Parfym & Kosmetika, Partihandel': 'Perfume & Cosmetics, Wholesale',
    'Organiska baskemikalier, tillverkning': 'Organic Basic Chemicals Manufacturing',
    'Optiska instrument & Fotoutrustning': 'Optical Instruments & Photographic Equipment',
    'Optiska fiberkabeltillverkning': 'Optical Fiber Cable Manufacturing',
    'Oorganiska baskemikalier, tillverkning': 'Inorganic Basic Chemicals Manufacturing',
    'Omsorg & Socialtjänst': 'Care & Social Services',
    'Nötkreatur & Bufflar, övriga': 'Other Cattle & Buffaloes',
    'Nöjes- & Temaparksverksamhet': 'Amusement & Theme Park Activities',
    'Näringslivsprogram, övriga': 'Business Programs, Other',
    'Nyhetsservice': 'News Service',
    'Naturvetenskaplig och Teknisk F&U': 'Natural Science and Technical R&D',
    'Möbler, Hushålls- & Järnhandelsvaror, Partihandel': 'Furniture, Household & Ironmongery, Wholesale',
    'Möbler & Heminredning, reparation': 'Furniture & Home Furnishing, Repair',
    'Musikinstrumenttillverkning': 'Musical Instrument Manufacturing',
    'Musik-, Dans- & Kulturell utbildning': 'Music, Dance & Cultural Education',
    'Museiverksamhet': 'Museum Activities',
    'Murbrukstillverkning': 'Mortar Manufacturing',
    'Motorfordonstillverkning': 'Motor Vehicle Manufacturing',
    'Motorer & Turbiner, tillverkning': 'Engines & Turbines, Manufacturing',
    'Motorcyklar, tillverkning': 'Motorcycle Manufacturing',
    'Motorbanor': 'Motor Racing Tracks',
    'Militärt försvar': 'Military Defense',
    'Militära stridsfordon, tillverkning': 'Military Armored Vehicle Manufacturing',
    'Metallvaror, reparation': 'Metal Goods, Repair',
    'Metaller & Metallmalmer, Partihandel': 'Metals & Metal Ores, Wholesale',
    'Metallavfall & Metallskrot, Partihandel': 'Metal Scrap & Metal Waste, Wholesale',
    'Mejerivarutillverkning': 'Dairy Product Manufacturing',
    'Mejeriprodukter, Ägg, Matolja & Matfett, Partihandel': 'Dairy Products, Eggs, Edible Oils & Fats,',
    'Medicin- & Dentalinstrumenttillverkning': 'Medical & Dental Instrument Manufacturing',
    'Mattor, tillverkning': 'Carpet Manufacturing',
    'Maskiner, Industriell utrustning, Partihandel': 'Machinery, Industrial Equipment, Wholesale',
    'Malmutvinning, övrig': 'Ore Mining, Other',
    'Magasinering & Varulagring': 'Warehousing & Storage',
    'Madrasser, tillverkning': 'Mattress Manufacturing',
    'Läder- & Skinnkläder, tillverkning': 'Leather & Fur Clothing Manufacturing',
    'Lufttransporter, stödtjänster': 'Air Transport Support Services',
    'Lufttransport, Passagerartrafik': 'Air Transport, Passenger Traffic',
    'Lufttransport, Godstrafik': 'Air Transport, Freight Traffic',
    'Luftfartyg & Rymdfarkoster, Reparation': 'Aircraft & Spacecraft, Repair',
    'Ljudinspelning & fonogramutgivning': 'Sound Recording & Phonogram Publishing',
    'Livsmedel, Dryck & Tobak,  Partihandel': 'Food, Beverage & Tobacco, Wholesale',
    'Linjebussverksamhet': 'Bus Transport Operations',
    'Lim, tillverkning': 'Adhesive Manufacturing',
    'Levande Djur, Partihandel': 'Live Animals, Wholesale',
    'Leasing av immateriell egendom & liknande prod.': 'Leasing of Intellectual Property & Similar Products',
    'Köksinredningar, tillverkning': 'Kitchen Furniture Manufacturing',
    'Kvarnprodukter': 'Mill Products',
    'Kultur, Miljö, Boende, administration': 'Culture, Environment, Accommodation,',
    'Kontorsutrustning & Datorer, Partihandel': 'Office Equipment & Computers, Wholesale',
    'Kontorstjänster': 'Office Services',
    'Kontorsmöbler, Partihandel': 'Office Furniture, Wholesale',
    'Kontorsmaskiner, tillverkning': 'Office Machine Manufacturing',
    'Kontorsmaskiner & Kontorsutrustning, Partihandel': 'Office Machines & Equipment, Wholesale',
    'Kontors- & Butiksmöber, tillverkning': 'Office & Shop Furniture, Manufacturing',
    'Konsumenttjänster, övriga': 'Consumer Services, Other',
    'Konstfiber, tillverkning': 'Artificial Fiber Manufacturing',
    'Kongresser & Mässor': 'Congresses & Fairs',
    'Kommunikationsutrustning, tillverkning': 'Communication Equipment Manufacturing',
    'Kommunikationsutrustning, reparation': 'Communication Equipment Repair',
    'Kollektivtrafik, övrig': 'Public Transport, Other',
    'Keramiska produkter, tillverkning': 'Ceramic Product Manufacturing',
    'Keramiska Golv- & Väggplattor': 'Ceramic Floor & Wall Tiles',
    'Kamel- & Kameldjursuppfödning': 'Camel & Camelid Farming',
    'Kalk & Gipstillverkning': 'Lime & Plaster Manufacturing',
    'Kaffe, Te, Kakao & Kryddor, Partihandel': 'Coffee, Tea, Cocoa & Spice,',
    'Kabeltillbehör, tillverkning': 'Cable Accessories Manufacturing',
    'Järnvägstransport-Godstrafik': 'Rail Transport-Freight Traffic',
    'Järnvägstransport- Passagerartrafik': 'Rail Transport-Passenger Traffic',
    'Juice & Safttillverkning': 'Juice & Nectar Production',
    'Jordbruks- & Textilråvaror, Provisionshandel': 'Agricultural & Textile Raw Materials, Wholesale',
    'Jord- & Skogsbruk, administration av program': 'Agriculture & Forestry, Program Administration',
    'Jakt': 'Hunting',
    'Intressorganisationer, övriga': 'Interest Organizations, Other',
    'Intressebev. Yrkesorg.': 'Interest Representation Prof. Org.',
    'Intressebev. Branschorg.': 'Interest Representation Industry Org.',
    'Intressebev. Arbetstagarorg.': 'Interest Representation Labor Org.',
    'Intressebev. Arbetsgivarorg.': 'Interest Representation Employer Org.',
    'Inspektion, Kontroll & Tillståndsgivning': 'Inspection, Control & Permitting',
    'Infrastrukturprogram': 'Infrastructure Programs',
    'Industrigasframställning': 'Industrial Gas Production',
    'Icke-farligt avfall': 'Non-Hazardous Waste',
    'Icke spec. handel med livsmedel, Partihandel': 'Non-Specific Food Trade, Wholesale',
    'Hästuppfödning': 'Horse Breeding',
    'Hälso- & Sjukvård, administration': 'Healthcare Administration',
    'Hushållsapparater, Hem & Trädgård, reparation': 'Household Appliances, Home & Garden, Repair',
    'Hudar, Skinn & Läder, Partihandel': 'Hides, Skins & Leather, Wholesale',
    'Hemelektronik, tillverkning': 'Consumer Electronics, Manufacturing',
    'Hemelektronik, reparation': 'Consumer Electronics, Repair',
    'Havs- & Sjöfart, Passagerartrafik': 'Maritime & Shipping, Passenger Traffic',
    'Havs- & Sjöfart, Godstrafik': 'Maritime & Shipping, Freight Traffic',
    'Hamngodshantering': 'Port Cargo Handling',
    'Gödsel- & Kväveprodukter, tillverkning': 'Fertilizer & Nitrogen Product Manufacturing',
    'Gymnasial utbildning': 'Upper Secondary Education',
    'Gymanläggningar': 'Gymnasiums',
    'Grönsaker, växthusodling': 'Vegetable Greenhouse Cultivation',
    'Grönsaker, frilandsodling': 'Vegetables, Open-Air Cultivation',
    'Grundskoleutbildning': 'Elementary Education',
    'Grundskole- & Gymnasieskoleutbildning': 'Elementary & Secondary School Education',
    'Grafisk produktion': 'Graphic Production',
    'Golfbanor': 'Golf Courses',
    'Godshantering, övrig': 'Cargo Handling, Other',
    'Glasfibertillverkning': 'Fiberglass Manufacturing',
    'Glas- & Glasvarutillverkning': 'Glass & Glassware Manufacturing',
    'Glas, Porslin & Rengöringsmedel, Partihandel': 'Glass, Porcelain & Cleaning Products, Wholesale',
    'Gjutning': 'Casting',
    'Gipsvarutillverkning': 'Gypsum Product Manufacturing',
    'Gashandel': 'Gas Trade',
    'Gasframställning': 'Gas Production',
    'Gasdistribution': 'Gas Distribution',
    'Garntillverkning': 'Yarn Manufacturing',
    'Förvärvsarbete i hushåll': 'Domestic Work',
    'Förskoleutbildning': 'Preschool Education',
    'Förlagsverksamhet, övrig': 'Publishing Activities, Other',
    'Företagstjänster, övriga': 'Business Services, Other',
    'Fönster av Trä, tillverkning': 'Wooden Window Manufacturing',
    'Får- & Getuppfödning': 'Sheep & Goat Farming',
    'Färgämnen, tillverkning': 'Dye Manufacturing',
    'Färg, Lack & Tryckfärg, tillverkning': 'Paint, Varnish & Printing Ink Manufacturing',
    'Fruktodling': 'Fruit Cultivation',
    'Frukt, Bär & Nötter, odling': 'Fruits, Berries & Nuts, Growing',
    'Fritids-& Nöjesverksamhet, övrig': 'Leisure & Entertainment Activities, Other',
    'Fotografiska & Optiska produkter, Partihandel': 'Photographic & Optical Products, Wholesale',
    'Folkhögskoleutbildning': 'Folk High School Education',
    'Fjäderfä, uppfödning': 'Poultry Farming',
    'Fisk, Skaldjur & Blötdjur, beredning': 'Fish, Seafood & Mollusk, Processing',
    'Filmvisning': 'Film Screening',
    'Film, Video & TV': 'Film, Video & TV',
    'Fiberväxtodling': 'Fiber Crop Cultivation',
    'Fibercementvarutillverkning': 'Fiber Cement Goods Manufacturing',
    'Fartyg & Båtar, Reparation': 'Ship & Boat Repair',
    'Farligt avfall': 'Hazardous Waste',
    'Fabriksblandad Betongtillverkning': 'Ready-Mixed Concrete Manufacturing',
    'Eteriska oljor, tillverkning': 'Essential Oils Manufacturing',
    'Emballage, Partihandel': 'Packaging, Wholesale',
    'Elöverföring': 'Electric Power Transmission',
    'Elhandel': 'Electricity Trading',
    'Elektronisk & Optisk utrustning, reparation': 'Electronic & Optical Equipment Repair',
    'Elektriska Hushållsmaskiner, tillverkning': 'Manufacture of Electric Household Appliances',
    'Eldistribution': 'Electric Distribution',
    'Eldfasta produkter': 'Refractory Products',
    'Elapparatur övrig, tillverkning': 'Other Electrical Equipment Manufacturing',
    'Eftergymnasial utbildning': 'Post-Secondary Education',
    'Dryckesframställning': 'Beverage Production',
    'Drycker, Partihandel': 'Beverages, Wholesale',
    'Domstolsverksamhet': 'Court Activities',
    'Djuruppfödning, övrig': 'Other Animal Farming',
    'Djurfoderframställning': 'Animal Feed Production',
    'Direktreklamverksamhet': 'Direct Mail Advertising',
    'Datoriserad materialhanteringsutr, Partihandel': 'Computerized Material Handling Equipment, Wholesale',
    'Datorer & Kringutrustning, tillverkning': 'Computers & Peripheral Equipment Manufacturing',
    'Datorer & Kringutrustning, reparation': 'Computers & Peripheral Equipment Repair',
    'Dagstidningsutgivning': 'Newspaper Publishing',
    'Cyklar & Invalidfordon, tillverkning': 'Bicycles & Invalid Vehicles Manufacturing',
    'Civilt försvar & Frivilligförsvar': 'Civil Defense & Voluntary Defense',
    'Choklad- & Konfektyrtillverkning': 'Chocolate & Confectionery Manufacturing',
    'Cementtillverkning': 'Cement Manufacturing',
    'Bärgning': 'Salvaging',
    'Byggmaterialtillverkning': 'Building Material Manufacturing',
    'Bud- & Kurirverksamhet': 'Courier & Messenger Activities',
    'Brand- & Räddningsverksamhet': 'Firefighting & Rescue Activities',
    'Borstbinderitillverkning': 'Brush and Broom Manufacturing',
    'Bokutgivning': 'Book Publishing',
    'Blommor & Växter, Partihandel': 'Flowers & Plants, Wholesale',
    'Blandat sortiment': 'Mixed Assortment',
    'Blandat jordbruk': 'Mixed Farming',
    'Bijouteritillverkning': 'Costume Jewelry Manufacturing',
    'Bibliotek': 'Library',
    'Betong-, Cement- & Gipsvaror, övriga': 'Concrete, Cement & Plaster Products, Other',
    'Belysningsarmaturtillverkning': 'Lighting Fixture Manufacturing',
    'Bekämpningsmedel & lantbrukskem, tillverkning': 'Pesticides & Agricultural Chemicals Manufacturing',
    'Begravningsverksamhet': 'Funeral Activities',
    'Batteri- och ackumulatortillverkning': 'Battery and Accumulator Manufacturing',
    'Basplast, tillverkning': 'Basic Plastic Manufacturing',
    'Band & Skivor, Partihandel': 'Tapes & Discs, Wholesale',
    'Bageri- & Mjölprodukter': 'Bakery & Flour Products',
    'Avloppsrening': 'Wastewater Treatment',
    'Avfall & Skrot, Partihandel': 'Waste & Scrap, Wholesale',
    'Arkivverksamhet': 'Archiving',
    'Arbetsmarknadsutbildning': 'Labor Market Training',
    'Arbets- & Skyddskläder, tillverkning': 'Workwear & Protective Clothing Manufacturing',
    'Annonstidningsutgivning': 'Free Newspaper Publishing'
            }

Max_rows = -1 # -1 = all
Earliest_year = 2012 # Note: the data before 2016 is unfortunately mostly incomplete.
Read_from_file = True
Raw_dataset_filename = FOLDER + CLEAN_DATASET_FILENAME
# Specify how many regions will be defined
N_lat_bins = 14
N_long_bins = 6
# If true, Export will allow to export category and region mappings, as well as the cleaned dataset to CSV files
Export = True

def Read_data(folder, max_rows, earliest_year):
    # This function looks for .csv files in the folder, which are expected to contain the raw web-scraped data.
    frames = [] # List of dataframes created out of each .csv
    earliest_year
    # Scan the folder to find all the files to be read
    df = pd.DataFrame()
    for file in os.listdir(folder):
        if len(df) >= max_rows and max_rows != -1:
            break
        else:
            if file.endswith('.csv'):
                file_path = os.path.join(folder, file)
                try:
                    # Read the content of the CSV and put it in a dataframe
                    print("Reading", file_path.split('\\')[-1])
                    df = pd.read_csv(file_path, encoding=ENCODING, sep=';', on_bad_lines='skip', low_memory=False)
                    frames.append(df)
                except Exception as e:
                    print(f"Error reading {file}: {e}")
    # Merge all the dataframes
    print("Concatenating dataframes...")
    concatenated_raw_data = pd.concat(frames, ignore_index=True)
    if max_rows == -1:
        return concatenated_raw_data
    elif max_rows >= 0:
        return concatenated_raw_data.head(max_rows)
    else:
        print("Max rows error")
        return concatenated_raw_data.head(0)

def Data_cleaning_1(df, earliest_year, sorted_columns_translation_table, activity_types_translation_table):
    # The purpose of this first cleaning is to clean the data to make it suitable for manual analysis.
    # The process doesn't alter the content of the raw data.
    # The data is processed only if the source isn't empty
    if len(df):
        # Remove the duplicates by using the organization number
        print("Dropping duplicates...")
        df = df.drop_duplicates(subset='orgnr')
        # Sorting the columns to have the KPIs first
        print("Sorting columns...")
        sorted_columns = [col for col in list(sorted_columns_translation_table.keys()) if col in df.columns]
        df = df[sorted_columns]

        # Rename the columns with english names
        print("Translating columns...")
        df.columns = list(sorted_columns_translation_table.values())

        # Translate the activity types in english
        print("Translating activity types...")
        df['category'] = df['category'].map(activity_types_translation_table)

        # Drop some columns that have little meaning
        drop_columns = ['orgnr', 'abv_hgrupp', 'linkTo', "Minoritetsintressen_2015", "Minoritetsintressen_2016", "Minoritetsintressen_2017", "Minoritetsintressen_2018", "Minoritetsintressen_2020"]
        columns_to_drop = set(df.columns).intersection(drop_columns)
        print("Dropping useless columns:", columns_to_drop)
        df = df.drop(columns=columns_to_drop)

        # Delete years columns that aren't requested
        years_to_drop = list(map(str, range(2011, earliest_year)))
        print("Dropping useless years:", years_to_drop)
        df = df.loc[:, ~df.columns.str.endswith(tuple(years_to_drop))]

        # Iterate over all columns and replace values only if they are strings and the column ends with a specific year
        # Some columns containing percentage values are converted to integer
        for col in df.columns:
            if any(year in col for year in map(str, range(2011, 2023))):
                print("Removing %, commas, spaces and converting to integers:", col)
                # Remove the percentage symbol, replace comma with dot, and spaces with an empty string;
                # finally, convert all values to integers. Note: percentage values are converted to integers taking into account the two decimal places present in the original data.
                df[col] = df[col].apply(lambda x: pd.to_numeric(str(x).replace(' ', '').rstrip('%').replace(',', '.'), errors='coerce'))
        return df

def Data_cleaning_2(df, unrelevant_features, lat_bins = 10, long_bins = 2):
    # This second cleaning process performs changes in the dataset.
    # This function alters the original features, to make them more suitable for machine learning.
    # Since the companies are uniquely identified by the new index, the juridical name is now useless and can be dropped
    # Note: the default values of lat_bins and long_bins are meant to roughly match the number of swedish counties and resembles Sweden's oblong shape as well.
    
    print("Dropping juridical name...")
    df = df.drop(["juridical name"], axis = 1)
    
    # The municipal seat is actually not very interesting, so it will be used only to fill the empty cells under "location", then dropped.
    print("Merging location and municipal seat...")
    df['location'] = df['location'].fillna(df["municipal seat"])
    df = df.drop(["municipal seat"], axis = 1)

    # Unfortunately there are several features that are mostly incomplete. For this reason, they are dropped rather than being completed
    unrelevant_features = [col for col in unrelevant_features if col in df.columns]
    print(f"Dropping unrelevant features: {unrelevant_features}")
    df = df.drop(unrelevant_features, axis = 1)

    # Status can be mapped to ordinals:
    print("Converting status to ordinals...")
    status_mapping = {"Bolaget är aktivt": 2, "Bolaget är inaktivt": 1, "Registrerad": -1} # -1 = only registered, probably recently bankrupted or liquidated
    df['Status'] = df['Status'].map(status_mapping)
    df = df.dropna(subset=['Status'])
    
    # The column with the registration date can be converted to an integer with just the year (month and day are not relevant for this dataset)
    print("Converting company registration date to integer...")
    df['registration year'] = pd.to_datetime(df['registration year'], errors='coerce').dt.year

    # The ownership can also be mapped to ordinals:
    print("Converting ownership to ordinals...")
    ownership_mapping = {"Privat, ej börsnoterat": 1, "-": -1} # 1 = not in stock exchange; -1 = deregistered; everything else becomes nan
    df['ownership'] = df['ownership'].map(ownership_mapping)
    df['ownership'] = df['ownership'].fillna(0)

    # The categories can be converted to ordinals
    print("Converting categories to ordinals...")
    category_mapping = {category: idx for idx, category in enumerate(df["category"].unique())}
    df["category"] = df["category"].map(category_mapping)

    # Managing the locations: the location feature can be replaced by two columns containing the corresponding coordinates.
    # This can be done with the help of another dataframe, called Coordinates.
    print("Replacing locations with coordinates...")
    df = pd.merge(df, Coordinates[['Location', 'Latitude', 'Longitude']], left_on='location', right_on='Location', how='left')
    df = df.drop(['location', 'Location'], axis=1)
    df = df.rename(columns={'Latitude': 'latitude', 'Longitude': 'longitude'})
    # For the purpose of machine learning, the exact coordinates aren't as meaningful as a coarse approximation can be
    # The latitude is split into lat_bins bands, while the longitude into long_bins bands.
    print("Grouping the coordinates into bins...")
    latitude_bins = pd.cut(df['latitude'], bins=lat_bins, labels=False, include_lowest=True)
    longitude_bins = pd.cut(df['longitude'], bins=long_bins, labels=False, include_lowest=True)
    df['latitude_bin'] = latitude_bins
    df['longitude_bin'] = longitude_bins
    df = df.drop(['latitude', 'longitude'], axis=1)
    # Convert latitude and longitude in a sort of univocal, scalar region code
    print("Converting the coordinates bins into region codes...")
    df['region_ID'] = df['latitude_bin'] * long_bins + df['longitude_bin']
    # Then the regional code is remapped in order to have a series without gaps
    region_ID_mapping = {region_ID: idx for idx, region_ID in enumerate(df["region_ID"].unique())}
    df['region_ID'] = df['region_ID'].map(region_ID_mapping)
    df = df.drop(['latitude_bin', 'longitude_bin'], axis=1)

    # Delete all the rows that contain too many NaNs. Before counting the NaNs, it's good to replace all those ones which are caused by the recent registration of a company.
    # Example: a company created in 2020 will have all the features related to earlier years set to NaN. This would inflate the count of NaNs illicitly.
    for col in df.columns:
        current_feature_split =col.split("_")
        if "20" in col:
            current_feature_year =  int(current_feature_split[-1])
        else:
            current_feature_year = 9999
        # The features dealing with years before the registration year are set to 0. In this way, all their NaNs are replaced,
        # while the NaNs in features with data from years after the registration year remain and will be handled later in the script.
        df[col] = np.where(df['registration year'] > current_feature_year, 0, df[col])
    # Now a new column with the count is created:
    print("Deleting hollow companies...")
    # Add a column (numbered 0) with the count of NaNs in each row.
    df = pd.concat([df, df.isna().sum(axis=1)], axis=1)
    # Then all the rows where NaNs are more than 33% the dataframe's width are deleted
    df = df[df[0] <= int(len(df.columns) * 0.33)]
    df = df.drop(0, axis=1)

    # Delete all the years with too much missing data
    print("Deleting hollow years...")
    for year in range(2012, 2023):
        missing_rate = Missing_data_rate(df, str(year))
        if missing_rate >= 50:
            columns_to_be_kept = df.filter(regex='^(?!.*_' + str(year) + '$)')
            df = df.loc[:, columns_to_be_kept.columns]

    # Remove "-" from the organization number, making it an integer
    print("Converting organization numbers to integers...")
    df.loc[:, "organization number"] = df["organization number"].str.replace("-", "").astype(np.int64)
    # Set the organization number as index
    print("Setting organization numbers as index...")
    df.set_index(df.columns[0], inplace=True)

    # Now it's time to fill all the empty data points. There are a bunch of criteria which are used according to a arbitrary hierarchy:
    # 1) All the data points dealing with years before the corresponding registration year are set to 0 (done))
    # 2) If the feature refers to a year and both the previous and the next years are available and have numbers, the replacement is the average of the two corresponding values in past and future.
    # 3) If only the previous year is available, its corresponding value is used as the replacement.
    # 4) If only the next year is available, its corresponding value is used as the replacement.
    # 5) If there isn't adjacent and valid data, the replacement is the median value calculated on the subset with the same category.
    # This is because the median values of each feature seem to greatly depend by the feature "category", so a fixed value wouldn't fit realistically as much as a specific value for certain subcategories.
    # Note: this insight has been obtained with help of Scan_stat_variation function defined later.
    # 6) If even the category median is a nan, the replacement is the general median over the whole current column.
    print("Replacing nans...")
    # The KPIs can be excluded from the column scan, since they can be calculated precisely out of the guessed source values instead of being guessed as well.
    # This is done just for a matter of data consistency.
    KPI_columns = pd.core.indexes.base.Index(['Net revenue per employee_2012',
                                                'Net revenue per employee_2013',
                                                'Net revenue per employee_2014',
                                                'Net revenue per employee_2015',
                                                'Net revenue per employee_2016',
                                                'Net revenue per employee_2017',
                                                'Net revenue per employee_2018',
                                                'Net revenue per employee_2019',
                                                'Net revenue per employee_2020',
                                                'Net revenue per employee_2021',
                                                'Net revenue per employee_2022',
                                                'Personnel costs per employee_2012',
                                                'Personnel costs per employee_2013',
                                                'Personnel costs per employee_2014',
                                                'Personnel costs per employee_2015',
                                                'Personnel costs per employee_2016',
                                                'Personnel costs per employee_2017',
                                                'Personnel costs per employee_2018',
                                                'Personnel costs per employee_2019',
                                                'Personnel costs per employee_2020',
                                                'Personnel costs per employee_2021',
                                                'Personnel costs per employee_2022',
                                                'EBITDA_2012',
                                                'EBITDA_2013',
                                                'EBITDA_2014',
                                                'EBITDA_2015',
                                                'EBITDA_2016',
                                                'EBITDA_2017',
                                                'EBITDA_2018',
                                                'EBITDA_2019',
                                                'EBITDA_2020',
                                                'EBITDA_2021',
                                                'EBITDA_2022',
                                                'Net revenue change_2012',
                                                'Net revenue change_2013',
                                                'Net revenue change_2014',
                                                'Net revenue change_2015',
                                                'Net revenue change_2016',
                                                'Net revenue change_2017',
                                                'Net revenue change_2018',
                                                'Net revenue change_2019',
                                                'Net revenue change_2020',
                                                'Net revenue change_2021',
                                                'Net revenue change_2022',
                                                'DuPont model_2012',
                                                'DuPont model_2013',
                                                'DuPont model_2014',
                                                'DuPont model_2015',
                                                'DuPont model_2016',
                                                'DuPont model_2017',
                                                'DuPont model_2018',
                                                'DuPont model_2019',
                                                'DuPont model_2020',
                                                'DuPont model_2021',
                                                'DuPont model_2022',
                                                'Profit margin_2012',
                                                'Profit margin_2013',
                                                'Profit margin_2014',
                                                'Profit margin_2015',
                                                'Profit margin_2016',
                                                'Profit margin_2017',
                                                'Profit margin_2018',
                                                'Profit margin_2019',
                                                'Profit margin_2020',
                                                'Profit margin_2021',
                                                'Profit margin_2022',
                                                'Gross profit margin_2012',
                                                'Gross profit margin_2013',
                                                'Gross profit margin_2014',
                                                'Gross profit margin_2015',
                                                'Gross profit margin_2016',
                                                'Gross profit margin_2017',
                                                'Gross profit margin_2018',
                                                'Gross profit margin_2019',
                                                'Gross profit margin_2020',
                                                'Gross profit margin_2021',
                                                'Gross profit margin_2022',
                                                'Working capital/revenue_2012',
                                                'Working capital/revenue_2013',
                                                'Working capital/revenue_2014',
                                                'Working capital/revenue_2015',
                                                'Working capital/revenue_2016',
                                                'Working capital/revenue_2017',
                                                'Working capital/revenue_2018',
                                                'Working capital/revenue_2019',
                                                'Working capital/revenue_2020',
                                                'Working capital/revenue_2021',
                                                'Working capital/revenue_2022',
                                                'Solvency_2012',
                                                'Solvency_2013',
                                                'Solvency_2014',
                                                'Solvency_2015',
                                                'Solvency_2016',
                                                'Solvency_2017',
                                                'Solvency_2018',
                                                'Solvency_2019',
                                                'Solvency_2020',
                                                'Solvency_2021',
                                                'Solvency_2022',
                                                'Quick ratio_2012',
                                                'Quick ratio_2013',
                                                'Quick ratio_2014',
                                                'Quick ratio_2015',
                                                'Quick ratio_2016',
                                                'Quick ratio_2017',
                                                'Quick ratio_2018',
                                                'Quick ratio_2019',
                                                'Quick ratio_2020',
                                                'Quick ratio_2021',
                                                'Quick ratio_2022'
                                                ])
    for col in df.columns.difference(KPI_columns):
        # Set current_feature_year according to the year contained in the current feature (column name)
        # previous_feature_year and next_feature_year are calculated consequently.
        current_feature_split = col.split("_")
        if "20" in col:
            current_feature_year = current_feature_split[-1]
            previous_feature_year = current_feature_split[0] + "_" + str(int(current_feature_year) - 1)
            next_feature_year = current_feature_split[0] + "_" + str(int(current_feature_year) + 1)
        else:
            current_feature_year = "9999"
            previous_feature_year = "9999"
            next_feature_year = "9999"
        # If the current feature year has both a predecessor and a successor, the nans are possibly replaced by the average of these two values
        if previous_feature_year in df.columns and next_feature_year in df.columns:
            # Both the previous and the next year may exist for the current feature; however, one or both of them may be nans as well.
            # This would mean to replace a nan with another nan.
            replace_values = (df[previous_feature_year] + df[next_feature_year]) / 2
            # Eventually either df[previous_feature_year] or df[next_feature_year] contain a nan, which causes a nan even in replace_values.
            # A simple way to replace even these residual nans is to replace them with the average over all the years of the same feature type:
            columns_to_average = [current_feature_split[0] + f"_{year}" for year in range(2012, 2023) if current_feature_split[0] + "_" + str(year) in df.columns]
            replace_values = replace_values.fillna(df[columns_to_average].mean(axis = 1))
            # Now that a series without nans has been obtained, it's time to use it to replace the nans in the main dataframe
            df[col] = df[col].fillna(replace_values)
        # Otherwise, the predecessor or the successor is used.
        elif previous_feature_year in df.columns:
            df[col] = df[col].fillna(df[previous_feature_year])
            # However, even this series can contain nans, so this must be handled.
            columns_to_average = [current_feature_split[0] + f"_{year}" for year in range(2012, 2023) if current_feature_split[0] + "_" + str(year) in df.columns]
            replace_values = df[columns_to_average].mean(axis = 1)
            df[col] = df[col].fillna(replace_values)
        elif next_feature_year in df.columns:
            df[col] = df[col].fillna(df[next_feature_year])
            # However, even this series can contain nans, so this must be handled: the mean value of the whole row is used.
            columns_to_average = [current_feature_split[0] + f"_{year}" for year in range(2012, 2023) if current_feature_split[0] + "_" + str(year) in df.columns]
            replace_values = df[columns_to_average].mean(axis = 1).fillna(0) # In the case of a whole row of nans, the values are replaced with zeros.
            df[col] = df[col].fillna(replace_values)
        elif "20" in col:
        # The case where no values in the closest years are available: the nans are here replaced by the median over the corresponding category
        # This is because the median values for certain features are quite specific for the category the company belongs to.
        # Note: this applies only to "year-features", so others like Status are skipped, since they all should have valid values already. 
            # Create a dictionary with the medians of the current column, calculated over each category
            category_median = df.groupby('category')[col].agg('median').to_dict()
            # Map the category with the corresponding medians
            mapped_category_median = df['category'].map(category_median)
            # The mapped series just defined can be used to fill the nans, since both df[col] and mapped_category_median have the same indexes.
            df[col] = df[col].fillna(mapped_category_median)
            # However, especially if the dataset is little and thus there isn't valid data in a whole category, the median itself can be a nan.
            # So, one more fill is needed, this time the general median value of the whole column.
            # In case the whole row is made of nans, then the median would be a nan as well, so these are replaced by zeros.
            df[col] = df[col].fillna(df[col].median()).fillna(0)
            # At this point, it would be still possible to have nans left, if the whole column the median is calculated over is made of nans.
            # However, such empty columns have already been dropped in the cleaning steps above.

    # After having replaced all the nans, the remaining ones in the KPI columns can be calculated in this second column scan:
    # As stated before, the features dealing with KPIs are calcutated out of other features; for consistency reasons, they can be calculated here,
    # instead of estimating their value with the same criteria used for filling the nans in the other features.
    KPI_df = pd.DataFrame()
    for col in KPI_columns:
        if col in df.columns:
            current_feature_year = col.split("_")[-1]
            # Each KPI is rebuilt by using the proper formulas
            print("Updating", col)
            if 'EBITDA' in col:
                # EBITDA is calculated as EBIT before depreciation and amortization. Here the values are not recalculated, except when they are nan.
                # Those nans are replaced with the corresponding EBIT
                KPI_df[col] = df['EBITDA_' + current_feature_year].fillna(df['EBIT_' + current_feature_year])
            elif 'Solvency' in col:
                # Infinite solvencies are replaced with zeros or hundreds, depending by the sign
                # Nans are due to years before the founding year, so the solvency is set to 100% by default, in such cases.
                KPI_df[col] = round(100 * (df['Equity_' + current_feature_year] + 0.786 * df['Untaxed reserves_' + current_feature_year]) / df['Liabilities and equity_' + current_feature_year], 2).replace(-np.inf, 0).replace(np.inf, 100).fillna(100.0)
            # The net revenue per employee might not be in the dataset because of it's often almost empty and so it's dropped previously by the algorithm
            elif 'Net revenue per employee' in col:
                # If there are no employees, the company is considered as it has one
                Current_employees = df['Number of employees_' + current_feature_year].replace({0: 1})
                KPI_df[col] = df['Net revenue_' + current_feature_year] / Current_employees
            elif 'Net revenue change_' in col:
                # Infinite changes are replaced by zeros;
                # if either the previous or the current net revenues aren't available, the column is set to 0.
                Current_net_revenue = 'Net revenue_' + current_feature_year
                Previous_net_revenue = 'Net revenue_' + str(int(current_feature_year) - 1)
                if Current_net_revenue in df.columns and Previous_net_revenue in df.columns:
                    KPI_df[col] = round((df[Current_net_revenue] / df[Previous_net_revenue] - 1) * 100, 2).replace([np.inf, -np.inf], 0).fillna(0.0)
                else:
                    KPI_df[col] = pd.Series([0] * len(df), name = col, index = df.index)
            elif 'DuPont model' in col:
                # The Dupont model is calculated as Profit margin x Net revenue, divided by the sum of all the assets; NaNs are replaced with zeros,
                # while -inf and inf are replaced with the double of the minimum value in the column and the maximum respectively.
                # This is done to avoid infinite values that can't be fed in certain machine learning algorhitms but, at the same time, provide meaningful replacements.
                Dupont_noinfs = df["Profit margin_" + current_feature_year] * df["Net revenue_" + current_feature_year] / (df["Subscribed unpaid capital_" + current_feature_year] + df["Fixed assets_" + current_feature_year] + df["Current assets_" + current_feature_year])
                Dupont_noinfs = Dupont_noinfs.replace([-np.inf, np.inf], np.nan).dropna()
                KPI_df[col] = round(df["Profit margin_" + current_feature_year] * df["Net revenue_" + current_feature_year] / (df["Subscribed unpaid capital_" + current_feature_year] + df["Fixed assets_" + current_feature_year] + df["Current assets_" + current_feature_year]), 2).replace(np.nan, 0).replace(-np.inf, 2 * min(Dupont_noinfs)).replace(np.inf, 2 * max(Dupont_noinfs))
            elif 'Profit margin' in col:
                # The source data seems to use mostly the EBIT to calculate the profit margin: 90% of the entries have a difference between generated and original within 1%.
                # However, sometimes (maybe due to errors in the balance sheet) the profit margin seems to be calculated by using the profit after financial items.
                KPI_df[col] = round(100 * df["EBIT_" + current_feature_year] / df["Net revenue_" + current_feature_year], 2).replace([np.nan, np.inf, -np.inf], 0)
            elif 'Gross profit margin' in col:
                # Nans in gross profit margin are usually due to lack of revenue and can't be calculated, so they are replaced with zeros
                KPI_df[col] = df[col].fillna(0.0)
            elif 'Working capital/revenue' in col:
                # The working capital is defined as Current assets minus Short-term liabilities, divided by the net revenue.
                # NaNs are replaced with zeros, while -inf and inf are replaced with the double of the minimum value in the column and the maximum respectively.
                # This is done to avoid infinite values that can't be fed in certain machine learning algorhitms but, at the same time, provide meaningful replacements.
                working_capital = df['Current assets_' + current_feature_year] - df['Short-term liabilities_' + current_feature_year]
                working_capital_revenue = 100 * working_capital / df['Net revenue_' + current_feature_year]
                working_capital_revenue_noinfs = working_capital_revenue.replace([-np.inf, np.inf], np.nan).dropna()
                max_working_capital_revenue = max(working_capital_revenue_noinfs)
                min_working_capital_revenue = min(working_capital_revenue_noinfs)
                KPI_df[col] = round(working_capital_revenue.replace(-np.inf, 2 * min_working_capital_revenue).replace(np.inf, 2 * max_working_capital_revenue), 2).replace(np.nan, 0)
            elif 'Quick ratio' in col:
                # If there are no short-term liabilities, the quick ratio is set conventionally to 100%
                KPI_df[col] = round((df['Current assets_' + current_feature_year] / df['Short-term liabilities_' + current_feature_year]).replace([np.inf, -np.inf, np.nan], 100), 2)
            elif 'Personnel costs per employee' in col:
                # The personnel costs' definition is company-dependent. It includes salary, social costs as well as other benefits that aren't available in this dataset.
                # The principle applied here is to leave the values as they are and only replace the nans. The nans are replaced, where possible, with the mean value over the row of personnel costs per employee.
                # If the mean personnel costs per employee is a nan as well (e. g. because all the items in the mean count are nans), it's replaced with 0.
                mean_personnel_costs_per_employee = df.filter(like = 'Personnel costs per employee').mean(axis = 1).fillna(0)
                KPI_df[col] = df['Personnel costs per employee_' + current_feature_year].fillna(mean_personnel_costs_per_employee)

            # else:
            #     print()
            # Those who got a nan, because of e. g. division by zero, are set to zero
    # Update the KPI columns in the main dataframe
    df.update(KPI_df)
    return df, category_mapping, region_ID_mapping, latitude_bins, longitude_bins

def Excel_with_template(df, filename):
    # Export to Excel by using a hard-coded file as a template
    try:
        template_path = 'Template.xlsx'
        template_wb = load_workbook(template_path)
        # Select or create a sheet in the template (replace 'Sheet1' with the name of your sheet)
        template_sheet = template_wb['Sheet1']
        # Add column names as headers in row 1
        for col_index, col_name in enumerate(df.columns, start=1):
            template_sheet.cell(row=1, column=col_index, value=col_name)
        # Get data from the DataFrame
        data = df.values.tolist()
        # Write data into the template starting from a certain cell
        for row_index, row_data in enumerate(data, start=2):
            for col_index, cell_value in enumerate(row_data, start=1):
                template_sheet.cell(row=row_index, column=col_index, value=cell_value)
        # Save the file with the new data
        template_wb.save(filename + '.xlsx')
        print("Dataset successfully exported to " + filename + ".xlsx.")
    except Exception as e:
        print(f"Failed to export the dataset to {filename}.xlsx. Error: {e}")
        try:
            df.to_csv(filename + '.csv', index=False)
            print("Dataset successfully exported to " + filename + ".csv")
        except:
            print(f"Failed to export the dataset to {filename}.csv.")

def Swedish_coordinates(df):
    # Retrieves a dataframe of location names and coordinates out of the main dataset
    locations = df["location"].drop_duplicates()
    latitudes = []
    longitudes = []
    i = 1
    length = len(locations)
    for location in locations:
        print(f"Getting {i} of {length}: {location}")
        coordinates = Get_coordinates(str(location) + ", Sweden")
        latitudes.append(coordinates[0])
        longitudes.append(coordinates[1])
        i += 1
    result = pd.DataFrame({"Location": locations, "Latitude": latitudes, "Longitude": longitudes})
    return result

def Get_coordinates(location_name):
    # Returns latitude and longitude, found with help of Nominatim
    geolocator = Nominatim(user_agent="my_geocoding_app")
    location = geolocator.geocode(location_name)
    if location:
        latitude, longitude = location.latitude, location.longitude
        return latitude, longitude
    else:
        print(f"Coordinates not found for {location_name}")
        return (-9999, -9999)

def Missing_data_rate(df, year):
    # This function computes the average rate of missing data (NaN) in the features of each year
    count_nan_per_column = df.isna().sum()
    nan_per_column_rate = 100 * count_nan_per_column / len(df)
    mean = nan_per_column_rate.filter(like=year).mean()
    return mean

def Group(df, feature, group, formula):
    # Return a dataframe containing the value (mean variance or whatever) of each category (value of a certain feature), together with the number of their occurrences
    # Values equal to -1 are filtered out.
    result = df[df[feature] != -1].groupby(group)[feature].agg([formula, 'count']).reset_index()
    result.rename(columns={formula: feature + '_' + formula, 'count': 'Occurrences'}, inplace=True)
    return result

def Scan_stat_variation(df, f1, f2, stat):
    # Scans through the columns and evaluates the coefficient of variation of a certain stat (e. g. median) in two features.
    # The function then prints wether first coefficient of variation is greater than the second one.
    # The aim of this function is to get an idea of which feature the stat may be calculated towards.
    # The reason behind this search is to find a good repalcement value for invalid data points.
    # Since a stat may vary much depending by other features (in this case region_ID or category),
    # it may be appropriate to replace the invalid data with a function of the feature with highest coefficient of variation, rather than a fixed value.
    count = 0
    for col in df.columns:
        a = Group(df, col, f1, stat)
        b = Group(df, col, f2, stat)
        a_cv = 100 * (a.std()[col + "_" + stat]) / a.mean()[col + "_" + stat]
        b_cv = 100 * (b.std()[col + "_" + stat]) / b.mean()[col + "_" + stat]
        count += 1 * (a_cv > b_cv)
        print(col, "\n", a_cv, b_cv, a_cv > b_cv, "\n")
    print(f"The variability of {f1} is {int(100 * count / len(df.columns))}% of the times greater than {f2}")




# MAIN

print(str(datetime.datetime.today().hour) + ":" + str(datetime.datetime.today().minute) + ":" + str(datetime.datetime.today().second))
if Read_from_file:
    print("Reading from", Raw_dataset_filename)
    if Max_rows != -1:
        Clean_data = pd.read_csv(Raw_dataset_filename, index_col=False, encoding=ENCODING).head(Max_rows)
    else:
        Clean_data = pd.read_csv(Raw_dataset_filename, index_col=False, encoding=ENCODING)
    Clean_data = Clean_data.drop(columns=["Unnamed: 0"])
else:
    print("Reading from folder", DATA_DIRECTORY)
    Raw_data = Read_data(DATA_DIRECTORY, Max_rows, Earliest_year)
    print(str(datetime.datetime.today().hour) + ":" + str(datetime.datetime.today().minute) + ":" + str(datetime.datetime.today().second))
    Clean_data = Data_cleaning_1(Raw_data, Earliest_year, Sorted_columns_translation_table, Activity_types_translation_table)
print(str(datetime.datetime.today().hour) + ":" + str(datetime.datetime.today().minute) + ":" + str(datetime.datetime.today().second))

Coordinates = pd.read_csv(FOLDER + COORDINATES_FILENAME, encoding=ENCODING, sep=",")
ML_data, Category_map, Region_map, Lat_bins, Long_bins = Data_cleaning_2(Clean_data, Unrelevant_features, N_lat_bins, N_long_bins)
if Export:
    pd.DataFrame(Category_map.items(), columns=['Category', 'Value']).to_csv(FOLDER + EXPORT_CATEGORY_MAP)
    pd.DataFrame(Category_map.items(), columns=['Region', 'Value']).to_csv(FOLDER + EXPORT_REGION_MAP)
    print("Writing to file...")
    if Max_rows != -1:
        ML_data.to_csv(f"{FOLDER}ML_data {N_lat_bins}x{N_long_bins} regions - {Max_rows} items.csv")
    else:
        ML_data.to_csv(f"{FOLDER}ML_data {N_lat_bins}x{N_long_bins} regions.csv")

print("Done!")